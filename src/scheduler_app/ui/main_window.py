from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QColor, QPainter, QTextCharFormat
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCalendarWidget,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QPlainTextEdit,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from scheduler_app.core.engine import SchedulerEngine
from scheduler_app.core.models import Assignment, POSITION_LABELS, SPECIAL_SL_MAIN_ROLES
from scheduler_app.data.repository import SchedulerRepository
from scheduler_app.services.exporter import export_csv, export_excel

try:
    import chinese_calendar
except Exception:
    chinese_calendar = None


RESULT_POSITIONS = ["SL_MAIN", "SL_AIR", "SL_GROUND", "TF_MAIN", "TF_GROUND", "FLEET_LEAD"]
RESULT_COL_POS = {3: "SL_MAIN", 4: "SL_AIR", 5: "SL_GROUND", 6: "TF_MAIN", 7: "TF_GROUND", 8: "FLEET_LEAD"}
WEEKDAY_MAP = {0: "\u5468\u4e00", 1: "\u5468\u4e8c", 2: "\u5468\u4e09", 3: "\u5468\u56db", 4: "\u5468\u4e94", 5: "\u5468\u516d", 6: "\u5468\u65e5"}


class PeopleTableWidget(QTableWidget):
    def __init__(self, on_reordered: Callable[[int, int], None], *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._on_reordered = on_reordered
        self._drag_allowed = False
        self._drag_row = -1

    def mousePressEvent(self, event):
        idx = self.indexAt(event.pos())
        self._drag_allowed = idx.isValid() and idx.column() in (0, 1)
        self._drag_row = idx.row() if self._drag_allowed else -1
        if self._drag_allowed:
            self.setCursor(Qt.ClosedHandCursor)
        super().mousePressEvent(event)

    def startDrag(self, supportedActions):
        if not self._drag_allowed:
            return
        super().startDrag(supportedActions)

    def dragEnterEvent(self, event):
        if event.source() is self:
            event.setDropAction(Qt.CopyAction)
            event.accept()
            return
        super().dragEnterEvent(event)

    def dragMoveEvent(self, event):
        if event.source() is self:
            event.setDropAction(Qt.CopyAction)
            event.accept()
            return
        super().dragMoveEvent(event)

    def dropEvent(self, event):
        if not self._drag_allowed or self._drag_row < 0:
            event.ignore()
            return
        idx = self.indexAt(event.pos())
        target_row = idx.row() if idx.isValid() else self.rowCount() - 1
        if idx.isValid():
            rect = self.visualRect(idx)
            if event.pos().y() > rect.center().y():
                target_row += 1
        if target_row > self.rowCount():
            target_row = self.rowCount()
        if target_row == self._drag_row or target_row == self._drag_row + 1:
            event.setDropAction(Qt.CopyAction)
            event.accept()
            self.setCursor(Qt.OpenHandCursor)
            return
        self._on_reordered(self._drag_row, target_row)
        self.setCursor(Qt.OpenHandCursor)
        event.setDropAction(Qt.CopyAction)
        event.accept()

    def mouseMoveEvent(self, event):
        idx = self.indexAt(event.pos())
        if idx.isValid() and idx.column() in (0, 1):
            self.setCursor(Qt.OpenHandCursor)
        else:
            self.setCursor(Qt.ArrowCursor)
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        idx = self.indexAt(event.pos())
        if idx.isValid() and idx.column() in (0, 1):
            self.setCursor(Qt.OpenHandCursor)
        else:
            self.setCursor(Qt.ArrowCursor)
        super().mouseReleaseEvent(event)

    def leaveEvent(self, event):
        self.setCursor(Qt.ArrowCursor)
        super().leaveEvent(event)


class PlanCalendarWidget(QCalendarWidget):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._manual_name_map: dict[date, str] = {}

    def set_manual_name_map(self, manual_name_map: dict[date, str]):
        self._manual_name_map = manual_name_map
        self.updateCells()

    def paintCell(self, painter: QPainter, rect, qdate):
        super().paintCell(painter, rect, qdate)
        name = self._manual_name_map.get(qdate.toPython())
        if not name:
            return
        painter.save()
        painter.setPen(QColor("#2b6de0"))
        f = painter.font()
        f.setPointSize(max(7, f.pointSize() - 1))
        painter.setFont(f)
        text_rect = rect.adjusted(1, rect.height() // 2 + 3, -1, -2)
        painter.drawText(text_rect, Qt.AlignHCenter | Qt.AlignBottom, name)
        painter.restore()


class LoginDialog(QDialog):
    def __init__(self, repo: SchedulerRepository, title: str, parent=None):
        super().__init__(parent)
        self.repo = repo
        self.account: dict | None = None
        self.setWindowTitle(title)
        self.setModal(True)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.username_edit = QLineEdit()
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)
        form.addRow("账号", self.username_edit)
        form.addRow("密码", self.password_edit)
        layout.addLayout(form)
        self.message_label = QLabel("")
        layout.addWidget(self.message_label)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._try_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _try_accept(self):
        account = self.repo.verify_login(self.username_edit.text(), self.password_edit.text())
        if account is None:
            self.message_label.setText("账号或密码错误")
            self.password_edit.selectAll()
            self.password_edit.setFocus()
            return
        self.account = account
        self.accept()


class AccountEditDialog(QDialog):
    def __init__(self, title: str, account: dict | None = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.username_edit = QLineEdit(account["username"] if account else "")
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)
        self.confirm_edit = QLineEdit()
        self.confirm_edit.setEchoMode(QLineEdit.Password)
        self.admin_check = QCheckBox("管理员账户")
        self.admin_check.setChecked(bool(account["is_admin"]) if account else False)
        form.addRow("账号", self.username_edit)
        form.addRow("密码", self.password_edit)
        form.addRow("确认密码", self.confirm_edit)
        form.addRow("", self.admin_check)
        layout.addLayout(form)
        self.hint_label = QLabel("编辑时密码留空表示不修改密码" if account else "")
        layout.addWidget(self.hint_label)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def payload(self) -> tuple[str, str, bool]:
        username = self.username_edit.text().strip()
        password = self.password_edit.text()
        confirm = self.confirm_edit.text()
        if not username:
            raise ValueError("用户名不能为空")
        if password != confirm:
            raise ValueError("两次输入的密码不一致")
        return username, password, self.admin_check.isChecked()


class MainWindow(QMainWindow):
    def __init__(self, db_path: Path, current_account: dict):
        super().__init__()
        self.setWindowTitle("\u6392\u73ed\u7cfb\u7edf")
        self.resize(1520, 940)
        self.repo = SchedulerRepository(db_path)
        self.repo.seed_if_empty()
        self.current_account = current_account
        self.current_assignments: list[Assignment] = []
        self.current_logs: list[str] = []
        self.employees = self.repo.load_employees()
        self.current_plan_date: date | None = None
        self.history_tail: list[dict] = []
        self.history_person_year_totals: dict[int, int] = {}
        self.history_team_year_totals: dict[str, int] = {}
        self.history_source_locked: bool = False
        self.history_source_type: str | None = None
        self.rerun_seed: int = 0
        self.last_non_account_tab_index = 0
        self.account_management_verified = False

        self._build_ui()
        self.refresh_people()
        self.refresh_leave_list()
        self.load_day_plan_table()

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        top = QHBoxLayout()
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDisplayFormat("yyyy-MM")
        self.start_date.setDate(date.today().replace(day=1))
        self.start_date.dateChanged.connect(self.on_start_date_changed)
        top.addWidget(QLabel("\u5e74\u6708"))
        top.addWidget(self.start_date)
        self.btn_reload_days = QPushButton("\u91cd\u5efa\u8ba1\u5212\u65e5\u5386")
        self.btn_reload_days.clicked.connect(self.load_day_plan_table)
        self.btn_generate = QPushButton("\u81ea\u52a8\u6392\u73ed")
        self.btn_generate.clicked.connect(self.generate_schedule)
        self.btn_export_csv = QPushButton("\u5bfc\u51faCSV")
        self.btn_export_csv.clicked.connect(self.on_export_csv)
        self.btn_export_xlsx = QPushButton("\u5bfc\u51faExcel")
        self.btn_export_xlsx.clicked.connect(self.on_export_xlsx)
        for b in [self.btn_reload_days, self.btn_generate, self.btn_export_csv, self.btn_export_xlsx]:
            top.addWidget(b)
        self.account_label = QLabel(
            f"当前账号：{self.current_account['username']}（{'管理员' if self.current_account['is_admin'] else '普通账号'}）"
        )
        top.addWidget(self.account_label)
        top.addStretch(1)
        layout.addLayout(top)

        self.tabs = QTabWidget()
        self.tabs.currentChanged.connect(self.on_tab_changed)
        layout.addWidget(self.tabs, 1)
        self.tab_plan = QWidget()
        self.tab_people = QWidget()
        self.tab_leave = QWidget()
        self.tab_result = QWidget()
        self.tab_summary = QWidget()
        self.tab_db = QWidget()
        self.tab_accounts = QWidget()
        self.tabs.addTab(self.tab_plan, "\u503c\u73ed\u8ba1\u5212")
        self.tabs.addTab(self.tab_people, "\u4eba\u5458\u7ba1\u7406")
        self.tabs.addTab(self.tab_leave, "\u8bf7\u5047\u7ba1\u7406")
        self.tabs.addTab(self.tab_result, "\u6392\u73ed\u7ed3\u679c")
        self.tabs.addTab(self.tab_summary, "\u7edf\u8ba1")
        self.tabs.addTab(self.tab_db, "\u6570\u636e\u5e93\u7ba1\u7406")
        self.tabs.addTab(self.tab_accounts, "账户管理")
        self._build_plan_tab()
        self._build_people_tab()
        self._build_leave_tab()
        self._build_result_tab()
        self._build_summary_tab()
        self._build_db_tab()
        self._build_accounts_tab()

    def _build_plan_tab(self):
        layout = QVBoxLayout(self.tab_plan)
        top = QHBoxLayout()
        left = QVBoxLayout()
        right = QVBoxLayout()
        self.plan_calendar = PlanCalendarWidget()
        self.plan_calendar.clicked.connect(self.on_plan_calendar_clicked)
        self.plan_calendar.setMinimumSize(760, 520)
        left.addWidget(QLabel("\u503c\u73ed\u8ba1\u5212\u65e5\u5386\uff08\u70b9\u51fb\u65e5\u671f\u8fdb\u884c\u914d\u7f6e\uff09"))
        btns = QHBoxLayout()
        self.btn_import_history = QPushButton("\u5bfc\u5165\u4e0a\u6708\u5386\u53f2\u6392\u73ed")
        self.btn_import_history.clicked.connect(self.import_history_from_file)
        self.btn_use_saved_history = QPushButton("\u5f15\u7528\u6392\u73ed\u6570\u636e\u5e93")
        self.btn_use_saved_history.clicked.connect(self.use_history_from_db)
        self.btn_clear_selected_leader = QPushButton("\u5220\u9664\u5df2\u9009\u503c\u73ed\u4eba\u5458")
        self.btn_clear_selected_leader.clicked.connect(self.clear_selected_plan_leader)
        self.btn_clear_saved_history = QPushButton("\u6e05\u9664\u5386\u53f2\u6570\u636e")
        self.btn_clear_saved_history.clicked.connect(self.clear_saved_history_data)
        btns.addWidget(self.btn_import_history)
        btns.addWidget(self.btn_use_saved_history)
        btns.addWidget(self.btn_clear_selected_leader)
        btns.addWidget(self.btn_clear_saved_history)
        left.addLayout(btns)
        left.addWidget(self.plan_calendar)
        self.plan_date_label = QLabel("\u5f53\u524d\u65e5\u671f: -")
        self.plan_default_label = QLabel("\u9ed8\u8ba4\u7c7b\u578b: -")
        right.addWidget(self.plan_date_label)
        right.addWidget(self.plan_default_label)
        right.addWidget(QLabel("\u65e5\u671f\u6807\u7b7e"))
        self.plan_tag_combo = QComboBox()
        self.plan_tag_combo.addItems(["\u5de5\u4f5c\u65e5", "\u7279\u6b8a\u65e5\u671f"])
        self.plan_tag_combo.currentTextChanged.connect(self.save_selected_plan_tag)
        right.addWidget(self.plan_tag_combo)
        right.addWidget(QLabel("\u7279\u6b8a\u65e5\u671f\u53cc\u6d41\u4e3b\u73ed\uff08\u4eba\u5de5\u6307\u5b9a\uff09"))
        self.plan_leader_combo = QComboBox()
        self.plan_leader_combo.currentIndexChanged.connect(self.save_selected_plan_leader)
        right.addWidget(self.plan_leader_combo)
        self.plan_day_tip = QLabel("\u63d0\u793a: \u7279\u6b8a\u65e5\u671f\u9700\u6307\u5b9a\u53cc\u6d41\u4e3b\u73ed\uff08\u5927\u961f\u957f/\u526f\u5927\u961f\u957f\uff09")
        right.addWidget(self.plan_day_tip)
        right.addStretch(1)
        top.addLayout(left, 3)
        top.addLayout(right, 1)
        layout.addLayout(top, 2)
        self.day_plan_table = QTableWidget(0, 4)
        self.day_plan_table.setHorizontalHeaderLabels(["\u65e5\u671f", "\u9ed8\u8ba4", "\u6807\u7b7e", "\u7279\u6b8a\u65e5\u53cc\u6d41\u4e3b\u73ed(\u4eba\u5de5)"])
        self.day_plan_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.day_plan_table.setMinimumHeight(360)
        layout.addWidget(self.day_plan_table, 4)
        self.log_box = QPlainTextEdit()
        self.log_box.setReadOnly(True)
        layout.addWidget(self.log_box, 1)

    def _build_people_tab(self):
        layout = QVBoxLayout(self.tab_people)
        form = QHBoxLayout()
        self.in_name = QLineEdit()
        self.in_team = QComboBox(); self.in_team.addItems(["\u4e8c", "\u4e09", "\u56db"])
        self.in_squad = QComboBox(); self.in_squad.addItems(["\u76f4\u5c5e", "\u4e00\u4e2d\u961f", "\u4e8c\u4e2d\u961f", "\u4e09\u4e2d\u961f", "\u56db\u4e2d\u961f"])
        self.in_group = QComboBox(); self.in_group.addItems(["\u7a7a\u52e4", "\u5730\u52e4"])
        self.in_role = QComboBox(); self.in_role.addItems(["\u5927\u961f\u957f", "\u526f\u5927\u961f\u957f_\u7a7a\u52e4", "\u526f\u5927\u961f\u957f_\u5730\u52e4", "\u515a\u603b\u652f\u4e66\u8bb0", "\u4e2d\u961f\u957f", "\u526f\u4e2d\u961f\u957f", "\u4e2d\u961f\u4e66\u8bb0"])
        self.in_participate = QCheckBox("\u53c2\u4e0e\u6392\u73ed"); self.in_participate.setChecked(True)
        self.btn_add_person = QPushButton("\u6dfb\u52a0\u4eba\u5458")
        self.btn_add_person.clicked.connect(self.add_person)
        self.btn_import_people = QPushButton("\u4eceExcel\u66ff\u6362\u4eba\u5458")
        self.btn_import_people.clicked.connect(self.import_people_from_excel)
        self.btn_export_people = QPushButton("\u5bfc\u51fa\u4eba\u5458")
        self.btn_export_people.clicked.connect(self.export_people_to_excel)
        for t, w in [("\u59d3\u540d", self.in_name), ("\u5927\u961f", self.in_team), ("\u4e2d\u961f", self.in_squad), ("\u7c7b\u522b", self.in_group), ("\u804c\u8d23", self.in_role)]:
            form.addWidget(QLabel(t)); form.addWidget(w)
        form.addWidget(self.in_participate)
        form.addWidget(self.btn_add_person)
        form.addWidget(self.btn_import_people)
        form.addWidget(self.btn_export_people)
        layout.addLayout(form)
        self.people_table = PeopleTableWidget(self.reorder_people_rows, 0, 8)
        self.people_table.setHorizontalHeaderLabels(["", "\u59d3\u540d", "\u5927\u961f", "\u4e2d\u961f", "\u7c7b\u522b", "\u804c\u8d23", "\u53c2\u4e0e\u6392\u73ed", "\u64cd\u4f5c"])
        self.people_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.people_table.setColumnWidth(0, 52)
        for col in range(1, 8):
            self.people_table.horizontalHeader().setSectionResizeMode(col, QHeaderView.Stretch)
        self.people_table.horizontalHeader().setStretchLastSection(True)
        self.people_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.people_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.people_table.setDragDropMode(QAbstractItemView.DragDrop)
        self.people_table.setDragEnabled(True)
        self.people_table.setAcceptDrops(True)
        self.people_table.setDropIndicatorShown(True)
        self.people_table.setDragDropOverwriteMode(False)
        self.people_table.setDefaultDropAction(Qt.CopyAction)
        self.people_table.horizontalHeader().setSectionsMovable(False)
        self.people_table.verticalHeader().setVisible(False)
        layout.addWidget(self.people_table)

    def _build_leave_tab(self):
        layout = QVBoxLayout(self.tab_leave)
        top = QHBoxLayout()
        self.leave_person = QComboBox()
        self.leave_person.currentIndexChanged.connect(self.refresh_leave_list)
        top.addWidget(QLabel("\u9009\u62e9\u4eba\u5458"))
        top.addWidget(self.leave_person)
        top.addWidget(QLabel("\u65e5\u5386\u70b9\u9009\u5207\u6362\u8bf7\u5047/\u53d6\u6d88\u8bf7\u5047"))
        self.btn_remove_selected_leave = QPushButton("\u5220\u9664\u9009\u4e2d\u8bf7\u5047")
        self.btn_remove_selected_leave.clicked.connect(self.remove_selected_leave)
        self.btn_clear_all_leave = QPushButton("\u5220\u9664\u8be5\u4eba\u5458\u5168\u90e8\u8bf7\u5047")
        self.btn_clear_all_leave.clicked.connect(self.clear_all_leave_for_person)
        self.btn_clear_everyone_leave = QPushButton("\u5220\u9664\u6240\u6709\u4eba\u5458\u6240\u6709\u8bf7\u5047")
        self.btn_clear_everyone_leave.clicked.connect(self.clear_all_leave_for_everyone)
        top.addWidget(self.btn_remove_selected_leave)
        top.addWidget(self.btn_clear_all_leave)
        top.addWidget(self.btn_clear_everyone_leave)
        top.addStretch(1)
        layout.addLayout(top)
        self.calendar_leave = QCalendarWidget()
        self.calendar_leave.clicked.connect(self.toggle_leave_on_calendar)
        layout.addWidget(self.calendar_leave, 3)
        self.leave_table = QTableWidget(0, 2)
        self.leave_table.setHorizontalHeaderLabels(["\u8bf7\u5047\u4fe1\u606f", "\u64cd\u4f5c"])
        self.leave_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.leave_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.leave_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.leave_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(self.leave_table, 2)
        leave_message_toolbar = QHBoxLayout()
        leave_message_toolbar.addWidget(QLabel("操作记录"))
        leave_message_toolbar.addStretch(1)
        self.btn_clear_leave_messages = QPushButton("清空记录")
        self.btn_clear_leave_messages.clicked.connect(self.clear_leave_messages)
        leave_message_toolbar.addWidget(self.btn_clear_leave_messages)
        layout.addLayout(leave_message_toolbar)
        self.leave_message_box = QPlainTextEdit()
        self.leave_message_box.setReadOnly(True)
        self.leave_message_box.setPlaceholderText("\u8bf7\u5047\u64cd\u4f5c\u63d0\u793a\u4f1a\u663e\u793a\u5728\u8fd9\u91cc")
        layout.addWidget(self.leave_message_box, 1)

    def _build_result_tab(self):
        layout = QVBoxLayout(self.tab_result)
        self.result_status_label = QLabel("\u7ed3\u679c\u72b6\u6001\uff1a\u672a\u751f\u6210")
        layout.addWidget(self.result_status_label)
        ops = QHBoxLayout()
        self.btn_rerun_with_mod = QPushButton("\u6839\u636e\u4fee\u6539\u91cd\u65b0\u6392\u73ed")
        self.btn_rerun_with_mod.clicked.connect(self.regenerate_with_overrides)
        self.btn_rerun_plan_only = QPushButton("\u5168\u90e8\u91cd\u6392")
        self.btn_rerun_plan_only.clicked.connect(self.regenerate_plan_only)
        self.btn_save_result = QPushButton("\u4fdd\u5b58\u6570\u636e")
        self.btn_save_result.clicked.connect(self.save_current_schedule_data)
        ops.addWidget(self.btn_rerun_with_mod)
        ops.addWidget(self.btn_rerun_plan_only)
        ops.addWidget(self.btn_save_result)
        ops.addStretch(1)
        layout.addLayout(ops)
        self.result_table = QTableWidget(0, 11)
        self.result_table.setHorizontalHeaderLabels(["\u65e5\u671f", "\u661f\u671f", "\u662f\u5426\u5468\u672b/\u8282\u5047\u65e5", "\u53cc\u6d41\u4e3b\u73ed", "\u53cc\u6d41\u526f\u73ed(\u7a7a\u52e4)", "\u53cc\u6d41\u526f\u73ed(\u5730\u52e4)", "\u5929\u5e9c\u4e3b\u73ed", "\u5929\u5e9c\u526f\u73ed", "\u673a\u961f\u603b\u8d1f\u8d23", "\u53cc\u6d41\u4e3b\u73ed\u5927\u961f", "\u5929\u5e9c\u4e3b\u73ed\u5927\u961f"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(self.result_table, 1)

    def _build_summary_tab(self):
        layout = QVBoxLayout(self.tab_summary)
        self.summary_table = QTableWidget(0, 5)
        self.summary_table.setHorizontalHeaderLabels(["\u59d3\u540d", "\u5927\u961f", "\u4e2d\u961f", "\u6708\u5ea6\u6b21\u6570", "\u5e74\u5ea6\u6b21\u6570(\u7d2f\u8ba1)"])
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.summary_table)

    def _build_db_tab(self):
        layout = QVBoxLayout(self.tab_db)
        layout.addWidget(QLabel("\u6570\u636e\u5e93\u7ba1\u7406\uff1a\u6bcf\u4e2a\u6708\u4efd\u53ea\u5bfc\u5165\u8be5\u6708\u672c\u6708\u6570\u636e\uff0c\u6267\u884c\u4e0b\u4e00\u4e2a\u6708\u6392\u73ed\u65f6\u4f1a\u81ea\u52a8\u8bfb\u53d6\u5b83\u4f5c\u4e3a\u4e0a\u6708\u6570\u636e\u3002"))
        self.db_table = QTableWidget(12, 4)
        self.db_table.setHorizontalHeaderLabels(["\u6708\u4efd", "\u72b6\u6001", "\u5bfc\u5165", "\u5220\u9664"])
        self.db_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.db_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.db_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.db_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.db_table.verticalHeader().setVisible(False)
        self.db_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.db_table.setSelectionMode(QAbstractItemView.NoSelection)
        layout.addWidget(self.db_table, 1)
        self.refresh_db_management()

    def _build_accounts_tab(self):
        layout = QVBoxLayout(self.tab_accounts)
        self.accounts_status_label = QLabel("账户管理需要管理员验证")
        layout.addWidget(self.accounts_status_label)
        top = QHBoxLayout()
        self.btn_add_account = QPushButton("新增账户")
        self.btn_add_account.clicked.connect(self.add_account)
        self.btn_add_account.setEnabled(False)
        top.addWidget(self.btn_add_account)
        self.btn_export_audit_logs = QPushButton("下载日志")
        self.btn_export_audit_logs.clicked.connect(self.export_audit_logs)
        self.btn_export_audit_logs.setEnabled(False)
        top.addWidget(self.btn_export_audit_logs)
        top.addStretch(1)
        layout.addLayout(top)
        layout.addWidget(QLabel("账户列表"))
        self.accounts_table = QTableWidget(0, 3)
        self.accounts_table.setHorizontalHeaderLabels(["用户名", "是否管理员账户", "操作"])
        self.accounts_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.accounts_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.accounts_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.accounts_table.verticalHeader().setVisible(False)
        self.accounts_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(self.accounts_table, 1)
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("操作日志"))
        filter_row.addWidget(QLabel("日志筛选账号"))
        self.audit_filter_combo = QComboBox()
        self.audit_filter_combo.currentIndexChanged.connect(self.refresh_audit_logs_table)
        filter_row.addWidget(self.audit_filter_combo)
        self.btn_refresh_audit_logs = QPushButton("刷新日志")
        self.btn_refresh_audit_logs.clicked.connect(self.refresh_audit_logs_view)
        self.btn_refresh_audit_logs.setEnabled(False)
        filter_row.addWidget(self.btn_refresh_audit_logs)
        filter_row.addStretch(1)
        layout.addLayout(filter_row)
        self.audit_logs_table = QTableWidget(0, 4)
        self.audit_logs_table.setHorizontalHeaderLabels(["时间", "账号", "操作", "详情"])
        self.audit_logs_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.audit_logs_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.audit_logs_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.audit_logs_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.audit_logs_table.verticalHeader().setVisible(False)
        self.audit_logs_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.audit_logs_table.setWordWrap(True)
        self.audit_logs_table.setAlternatingRowColors(True)
        self.audit_logs_table.verticalHeader().setDefaultSectionSize(44)
        layout.addWidget(self.audit_logs_table, 1)

    def on_tab_changed(self, index: int):
        if not hasattr(self, "tab_accounts"):
            return
        if self.tabs.widget(index) is self.tab_accounts:
            if not self.current_account["is_admin"]:
                QMessageBox.warning(self, "无权限", "普通账号不能进入账户管理")
                self.tabs.blockSignals(True)
                self.tabs.setCurrentIndex(self.last_non_account_tab_index)
                self.tabs.blockSignals(False)
                return
            if not self.prompt_admin_access():
                self.tabs.blockSignals(True)
                self.tabs.setCurrentIndex(self.last_non_account_tab_index)
                self.tabs.blockSignals(False)
                return
            self.account_management_verified = True
            self.btn_add_account.setEnabled(True)
            self.btn_refresh_audit_logs.setEnabled(True)
            self.btn_export_audit_logs.setEnabled(True)
            self.refresh_accounts_table()
            self.refresh_audit_filter_combo()
            self.refresh_audit_logs_table()
            self.accounts_status_label.setText("账户管理：已通过管理员验证")
            return
        self.last_non_account_tab_index = index
        self.account_management_verified = False
        if hasattr(self, "accounts_status_label"):
            self.accounts_status_label.setText("账户管理需要管理员验证")
        if hasattr(self, "btn_add_account"):
            self.btn_add_account.setEnabled(False)
        if hasattr(self, "btn_refresh_audit_logs"):
            self.btn_refresh_audit_logs.setEnabled(False)
        if hasattr(self, "btn_export_audit_logs"):
            self.btn_export_audit_logs.setEnabled(False)
        if hasattr(self, "accounts_table"):
            self.accounts_table.setRowCount(0)
        if hasattr(self, "audit_filter_combo"):
            self.audit_filter_combo.blockSignals(True)
            self.audit_filter_combo.clear()
            self.audit_filter_combo.blockSignals(False)
        if hasattr(self, "audit_logs_table"):
            self.audit_logs_table.setRowCount(0)

    def prompt_admin_access(self) -> bool:
        dlg = LoginDialog(self.repo, "管理员验证", self)
        dlg.password_edit.setPlaceholderText("请输入管理员密码")
        if dlg.exec() != QDialog.Accepted:
            return False
        if not dlg.account or not dlg.account["is_admin"]:
            QMessageBox.warning(self, "无权限", "仅管理员账号可以进入账户管理")
            return False
        return True

    def _audit(self, action_type: str, action_label: str, details_text: str) -> None:
        self.repo.add_audit_log(
            self.current_account["id"],
            self.current_account["username"],
            action_type,
            action_label,
            details_text,
        )

    def refresh_accounts_table(self):
        accounts = self.repo.list_accounts()
        self.accounts_table.setRowCount(len(accounts))
        for row, account in enumerate(accounts):
            self.accounts_table.setItem(row, 0, QTableWidgetItem(account["username"]))
            self.accounts_table.setItem(row, 1, QTableWidgetItem("是" if account["is_admin"] else "否"))
            ops = QWidget()
            ops_layout = QHBoxLayout(ops)
            ops_layout.setContentsMargins(0, 0, 0, 0)
            btn_edit = QPushButton("修改")
            btn_edit.clicked.connect(lambda _, a=account: self.edit_account(a))
            btn_delete = QPushButton("删除")
            btn_delete.clicked.connect(lambda _, a=account: self.delete_account(a))
            ops_layout.addWidget(btn_edit)
            ops_layout.addWidget(btn_delete)
            self.accounts_table.setCellWidget(row, 2, ops)

    def refresh_audit_filter_combo(self):
        self.audit_filter_combo.blockSignals(True)
        self.audit_filter_combo.clear()
        self.audit_filter_combo.addItem("全部账号", None)
        seen: set[tuple[int, str]] = set()
        for account in self.repo.list_accounts():
            key = (int(account["id"]), str(account["username"]))
            if key in seen:
                continue
            seen.add(key)
            self.audit_filter_combo.addItem(account["username"], account["id"])
        for entry in self.repo.list_audit_accounts():
            key = (int(entry["account_id"]), str(entry["username_snapshot"]))
            if key in seen:
                continue
            seen.add(key)
            self.audit_filter_combo.addItem(entry["username_snapshot"], entry["account_id"])
        self.audit_filter_combo.blockSignals(False)

    def refresh_audit_logs_table(self):
        if not self.account_management_verified:
            self.audit_logs_table.setRowCount(0)
            return
        account_id = self.audit_filter_combo.currentData()
        logs = self.repo.load_audit_logs(account_id)
        self.audit_logs_table.setRowCount(len(logs))
        for row, log in enumerate(logs):
            self.audit_logs_table.setItem(row, 0, QTableWidgetItem(log["created_at"]))
            self.audit_logs_table.setItem(row, 1, QTableWidgetItem(log["username_snapshot"]))
            self.audit_logs_table.setItem(row, 2, QTableWidgetItem(log["action_label"]))
            detail_item = QTableWidgetItem(log["details_text"])
            detail_item.setToolTip(log["details_text"])
            self.audit_logs_table.setItem(row, 3, detail_item)
            self.audit_logs_table.setRowHeight(row, 52)

    def refresh_audit_logs_view(self):
        if not self.account_management_verified:
            return
        self.refresh_audit_filter_combo()
        self.refresh_audit_logs_table()

    def export_audit_logs(self):
        if not self.account_management_verified:
            return
        account_id = self.audit_filter_combo.currentData()
        logs = self.repo.load_audit_logs(account_id)
        default_name = f"audit_logs_{date.today().isoformat()}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "下载日志", default_name, "Excel (*.xlsx)")
        if not file_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "操作日志"
            ws.append(["时间", "账号", "操作", "详情"])
            for log in logs:
                ws.append([log["created_at"], log["username_snapshot"], log["action_label"], log["details_text"]])
            for col, width in {"A": 22, "B": 16, "C": 18, "D": 60}.items():
                ws.column_dimensions[col].width = width
            wb.save(file_path)
            QMessageBox.information(self, "导出完成", f"已导出 {len(logs)} 条日志到：\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def add_account(self):
        if not self.account_management_verified:
            return
        dlg = AccountEditDialog("新增账户", parent=self)
        if dlg.exec() != QDialog.Accepted:
            return
        try:
            username, password, is_admin = dlg.payload()
            if not password:
                raise ValueError("密码不能为空")
            self.repo.create_account(username, password, is_admin)
            self._audit("account_create", "新增账户", f"新增账户：{username}（{'管理员' if is_admin else '普通账号'}）")
            self.refresh_accounts_table()
            self.refresh_audit_filter_combo()
            self.refresh_audit_logs_table()
        except Exception as e:
            QMessageBox.critical(self, "新增失败", str(e))

    def edit_account(self, account: dict):
        if not self.account_management_verified:
            return
        dlg = AccountEditDialog("修改账户", account=account, parent=self)
        if dlg.exec() != QDialog.Accepted:
            return
        try:
            username, password, is_admin = dlg.payload()
            old_username = account["username"]
            old_role = "管理员" if account["is_admin"] else "普通账号"
            self.repo.update_account(account["id"], username, password or None, is_admin)
            if account["id"] == self.current_account["id"]:
                self.current_account["username"] = username
                self.current_account["is_admin"] = is_admin
                self.account_label.setText(
                    f"当前账号：{self.current_account['username']}（{'管理员' if self.current_account['is_admin'] else '普通账号'}）"
                )
            self._audit(
                "account_update",
                "修改账户",
                f"修改账户：{old_username} -> {username}，权限：{old_role} -> {'管理员' if is_admin else '普通账号'}"
                + ("，已修改密码" if password else ""),
            )
            self.refresh_accounts_table()
            self.refresh_audit_filter_combo()
            self.refresh_audit_logs_table()
        except Exception as e:
            QMessageBox.critical(self, "修改失败", str(e))

    def delete_account(self, account: dict):
        if not self.account_management_verified:
            return
        if account["id"] == self.current_account["id"]:
            QMessageBox.warning(self, "删除失败", "不能删除当前已登录账号")
            return
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定删除账户 {account['username']} 吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        try:
            self.repo.delete_account(account["id"])
            self._audit("account_delete", "删除账户", f"删除账户：{account['username']}")
            self.refresh_accounts_table()
            self.refresh_audit_filter_combo()
            self.refresh_audit_logs_table()
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e))

    def on_start_date_changed(self, *_):
        self.history_source_locked = False
        self.history_source_type = None
        current = self.start_date.date().toPython().replace(day=1)
        self.start_date.blockSignals(True)
        self.start_date.setDate(current)
        self.start_date.blockSignals(False)
        if hasattr(self, "plan_calendar"):
            self.plan_calendar.setSelectedDate(QDate(current.year, current.month, 1))
            self.plan_calendar.setCurrentPage(current.year, current.month)
        self.current_plan_date = current
        self.load_day_plan_table()
        self.refresh_db_management()

    def _month_start(self) -> date:
        return self.start_date.date().toPython().replace(day=1)

    def _month_end(self) -> date:
        start = self._month_start()
        next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
        return next_month - timedelta(days=1)

    def refresh_db_management(self):
        if not hasattr(self, "db_table"):
            return
        year = self._month_start().year
        saved_months = self.repo.list_saved_months(f"{year:04d}")
        self.db_table.setRowCount(12)
        for month in range(1, 13):
            row = month - 1
            month_key = f"{year:04d}-{month:02d}"
            has_data = month_key in saved_months
            self.db_table.setRowHeight(row, 40)
            self.db_table.setItem(row, 0, QTableWidgetItem(month_key))
            self.db_table.setItem(row, 1, QTableWidgetItem("\u6709\u6570\u636e" if has_data else "\u65e0\u6570\u636e"))
            btn_import = QPushButton("\u8986\u76d6\u5bfc\u5165\u5f53\u6708\u6570\u636e" if has_data else "\u5bfc\u5165\u5f53\u6708\u6570\u636e")
            btn_import.setEnabled(True)
            btn_import.setMinimumHeight(30)
            btn_import.clicked.connect(lambda _, mk=month_key: self.import_db_month(mk))
            self.db_table.setCellWidget(row, 2, btn_import)
            btn_delete = QPushButton("\u5220\u9664")
            btn_delete.setEnabled(True)
            btn_delete.setMinimumHeight(30)
            btn_delete.clicked.connect(lambda _, mk=month_key: self.delete_db_month(mk))
            self.db_table.setCellWidget(row, 3, btn_delete)

    def _default_tag(self, d: date) -> str:
        if chinese_calendar is not None:
            try:
                if chinese_calendar.is_holiday(d) or d.weekday() >= 5:
                    return "SPECIAL"
                return "WORKDAY"
            except Exception:
                pass
        return "SPECIAL" if d.weekday() >= 5 else "WORKDAY"

    def _paint_plan_calendar(self):
        start = self._month_start()
        end = self._month_end()
        manual_tags = self.repo.load_day_tags()
        fmt_reset = QTextCharFormat()
        fmt_weekend = QTextCharFormat()
        fmt_weekend.setForeground(QColor("#ff0000"))
        fmt_manual_special = QTextCharFormat()
        fmt_manual_special.setForeground(QColor("#ff0000"))
        fmt_manual_workday = QTextCharFormat()
        fmt_manual_workday.setForeground(QColor("#000000"))
        d = start
        while d <= end:
            qd = QDate.fromString(d.isoformat(), "yyyy-MM-dd")
            self.plan_calendar.setDateTextFormat(qd, fmt_reset)
            d += timedelta(days=1)
        d = start
        while d <= end:
            default_tag = self._default_tag(d)
            tagged = manual_tags.get(d)
            qd = QDate.fromString(d.isoformat(), "yyyy-MM-dd")
            if tagged is not None and tagged != default_tag:
                self.plan_calendar.setDateTextFormat(qd, fmt_manual_special if tagged == "SPECIAL" else fmt_manual_workday)
            elif d.weekday() >= 5:
                self.plan_calendar.setDateTextFormat(qd, fmt_weekend)
            d += timedelta(days=1)

    def refresh_people(self):
        self.employees = self.repo.load_employees()
        self.people_table.setRowCount(len(self.employees))
        self.leave_person.clear()
        self.people_table.blockSignals(True)
        for i, e in enumerate(self.employees):
            order_item = QTableWidgetItem(str(i + 1))
            order_item.setTextAlignment(Qt.AlignCenter)
            order_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsDragEnabled)
            order_item.setData(Qt.UserRole, e.id)
            self.people_table.setItem(i, 0, order_item)
            name_item = QTableWidgetItem(e.name)
            name_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsDragEnabled)
            name_item.setData(Qt.UserRole, e.id)
            self.people_table.setItem(i, 1, name_item)
            cb_team = QComboBox(); cb_team.addItems(["二", "三", "四"]); cb_team.setCurrentText(e.team)
            cb_squad = QComboBox(); cb_squad.addItems(["直属", "一中队", "二中队", "三中队", "四中队"]); cb_squad.setCurrentText(e.squad)
            cb_group = QComboBox(); cb_group.addItems(["空勤", "地勤"]); cb_group.setCurrentText(e.duty_group)
            cb_role = QComboBox(); cb_role.addItems(["大队长", "副大队长_空勤", "副大队长_地勤", "党总支书记", "中队长", "副中队长", "中队书记"]); cb_role.setCurrentText(e.role)
            cb_team.currentTextChanged.connect(lambda _, eid=e.id, a=cb_team, b=cb_squad, c=cb_group, d=cb_role: self.update_person_meta(eid, a, b, c, d))
            cb_squad.currentTextChanged.connect(lambda _, eid=e.id, a=cb_team, b=cb_squad, c=cb_group, d=cb_role: self.update_person_meta(eid, a, b, c, d))
            cb_group.currentTextChanged.connect(lambda _, eid=e.id, a=cb_team, b=cb_squad, c=cb_group, d=cb_role: self.update_person_meta(eid, a, b, c, d))
            cb_role.currentTextChanged.connect(lambda _, eid=e.id, a=cb_team, b=cb_squad, c=cb_group, d=cb_role: self.update_person_meta(eid, a, b, c, d))
            self.people_table.setCellWidget(i, 2, cb_team)
            self.people_table.setCellWidget(i, 3, cb_squad)
            self.people_table.setCellWidget(i, 4, cb_group)
            self.people_table.setCellWidget(i, 5, cb_role)
            ck = QCheckBox(); ck.setChecked(e.participate)
            ck.stateChanged.connect(lambda state, eid=e.id: self.update_participate_state(eid, bool(state)))
            self.people_table.setCellWidget(i, 6, ck)
            btn = QPushButton("\u5220\u9664")
            btn.clicked.connect(lambda _, eid=e.id: self.delete_person(eid))
            self.people_table.setCellWidget(i, 7, btn)
            self.leave_person.addItem(f"{e.id}-{e.name}-{e.team}-{e.squad}", e.id)
        self.people_table.blockSignals(False)
        self._reload_plan_day_leaders()

    def update_person_meta(self, employee_id: int, team_box: QComboBox, squad_box: QComboBox, group_box: QComboBox, role_box: QComboBox):
        before = {e.id: e for e in self.employees}.get(employee_id)
        self.repo.update_employee_meta(employee_id, team_box.currentText(), squad_box.currentText(), group_box.currentText(), role_box.currentText())
        self.employees = self.repo.load_employees()
        after = {e.id: e for e in self.employees}.get(employee_id)
        if before is not None and after is not None:
            self._audit(
                "person_update",
                "修改人员信息",
                f"{before.name}：{before.team}-{before.squad}-{before.duty_group}-{before.role} -> "
                f"{after.team}-{after.squad}-{after.duty_group}-{after.role}",
            )
        self._reload_plan_day_leaders()

    def update_participate_state(self, employee_id: int, participate: bool):
        employee = {e.id: e for e in self.employees}.get(employee_id)
        self.repo.set_participate(employee_id, participate)
        if employee is not None:
            self._audit(
                "person_participate",
                "修改参与排班",
                f"{employee.name}：{'参与排班' if participate else '不参与排班'}",
            )

    def persist_people_order(self):
        ordered_ids = []
        for i in range(self.people_table.rowCount()):
            item = self.people_table.item(i, 0)
            if item is None:
                item = self.people_table.item(i, 1)
            if item is not None:
                eid = item.data(Qt.UserRole)
                if eid is not None:
                    ordered_ids.append(int(eid))
        if ordered_ids:
            self.repo.reorder_employees(ordered_ids)

    def reorder_people_rows(self, source_row: int, target_row: int):
        if source_row < 0 or source_row >= len(self.employees):
            return
        ordered_ids = [e.id for e in self.employees]
        moved_id = ordered_ids.pop(source_row)
        insert_row = max(0, min(target_row, len(ordered_ids)))
        if insert_row > source_row:
            insert_row -= 1
        ordered_ids.insert(insert_row, moved_id)
        moved_employee = {e.id: e for e in self.employees}.get(moved_id)
        self.repo.reorder_employees(ordered_ids)
        if moved_employee is not None:
            self._audit("person_reorder", "调整人员顺序", f"调整人员顺序：{moved_employee.name} 移动到第 {insert_row + 1} 位")
        self.refresh_people()

    def _reload_plan_day_leaders(self):
        self.plan_leader_combo.blockSignals(True)
        self.plan_leader_combo.clear()
        self.plan_leader_combo.addItem("(\u7a7a)", None)
        for e in [x for x in self.employees if x.role in SPECIAL_SL_MAIN_ROLES]:
            self.plan_leader_combo.addItem(f"{e.name}-{e.team}-{e.squad}-{e.role}", e.id)
        self.plan_leader_combo.blockSignals(False)

    def add_person(self):
        name = self.in_name.text().strip()
        if not name:
            QMessageBox.warning(self, "\u63d0\u793a", "\u8bf7\u8f93\u5165\u59d3\u540d")
            return
        self.repo.add_employee(name, self.in_team.currentText(), self.in_squad.currentText(), self.in_group.currentText(), self.in_role.currentText(), self.in_participate.isChecked())
        self._audit(
            "person_add",
            "新增人员",
            f"新增人员：{name}（{self.in_team.currentText()}-{self.in_squad.currentText()}，{self.in_group.currentText()}，{self.in_role.currentText()}）",
        )
        self.in_name.clear()
        self.refresh_people()
        self.load_day_plan_table()

    def import_people_from_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "\u9009\u62e9\u4eba\u5458Excel\u6587\u4ef6", "", "Excel (*.xlsx *.xlsm)")
        if not file_path:
            return
        try:
            count = self.repo.replace_employees_from_xlsx(Path(file_path))
            self.refresh_people()
            self.refresh_leave_list()
            self.load_day_plan_table()
            self._audit("people_import", "替换人员", f"从 Excel 替换人员，共导入 {count} 条，文件：{Path(file_path).name}")
            QMessageBox.information(self, "\u5bfc\u5165\u5b8c\u6210", f"\u5df2\u66ff\u6362\u4eba\u5458\uff0c\u5171\u5bfc\u5165 {count} \u6761\u3002")
        except Exception as e:
            QMessageBox.critical(self, "\u5bfc\u5165\u5931\u8d25", str(e))

    def delete_person(self, eid: int):
        by_id = {e.id: e for e in self.employees}
        self.repo.delete_employee(eid)
        if eid in by_id:
            e = by_id[eid]
            self._audit("person_delete", "删除人员", f"删除人员：{e.name}（{e.team}-{e.squad}）")
        self.refresh_people()
        self.load_day_plan_table()

    def toggle_leave_on_calendar(self):
        eid = self.leave_person.currentData()
        if eid is None:
            return
        d = self.calendar_leave.selectedDate().toPython()
        is_leave = self.repo.toggle_leave(eid, d)
        person_label = self.leave_person.currentText().split("-", 2)[1] if "-" in self.leave_person.currentText() else self.leave_person.currentText()
        action = "\u65b0\u589e\u8bf7\u5047" if is_leave else "\u53d6\u6d88\u8bf7\u5047"
        self._audit("leave_toggle", "修改请假", f"{person_label}：{d.isoformat()}，{action}")
        self.leave_message_box.appendPlainText(f"{person_label}\uff1a{d.year}/{d.month}/{d.day}\uff0c{action}")
        self.refresh_leave_list()

    def refresh_leave_list(self):
        self.leave_table.setRowCount(0)
        eid = self.leave_person.currentData()
        if eid is None:
            return
        person_label = self.leave_person.currentText().split("-", 2)[1] if "-" in self.leave_person.currentText() else self.leave_person.currentText()
        dates = sorted(dd for pid, dd in self.repo.load_leaves() if pid == eid)
        self.leave_table.setRowCount(len(dates))
        for i, d in enumerate(dates):
            item = QTableWidgetItem(f"{person_label}：{d.year}/{d.month}/{d.day}")
            item.setData(Qt.UserRole, d.isoformat())
            self.leave_table.setItem(i, 0, item)
            btn = QPushButton("删除")
            btn.clicked.connect(lambda _, dd=d, ee=eid: self.remove_leave_item(ee, dd))
            self.leave_table.setCellWidget(i, 1, btn)

    def remove_selected_leave(self):
        eid = self.leave_person.currentData()
        row = self.leave_table.currentRow()
        if eid is None or row < 0:
            return
        d = date.fromisoformat(self.leave_table.item(row, 0).data(Qt.UserRole))
        self.repo.remove_leave(eid, d)
        person_label = self.leave_person.currentText().split("-", 2)[1] if "-" in self.leave_person.currentText() else self.leave_person.currentText()
        self._audit("leave_remove", "删除请假", f"{person_label}：{d.isoformat()}")
        self.leave_message_box.appendPlainText(f"{person_label}：{d.year}/{d.month}/{d.day}，删除请假")
        self.refresh_leave_list()

    def remove_leave_item(self, eid: int, d: date):
        self.repo.remove_leave(eid, d)
        employee = {e.id: e for e in self.employees}.get(eid)
        if employee is not None:
            self._audit("leave_remove", "删除请假", f"{employee.name}：{d.isoformat()}")
            self.leave_message_box.appendPlainText(f"{employee.name}：{d.year}/{d.month}/{d.day}，删除请假")
        self.refresh_leave_list()

    def clear_all_leave_for_person(self):
        eid = self.leave_person.currentData()
        if eid is None:
            return
        self.repo.clear_leaves_for_employee(eid)
        employee = {e.id: e for e in self.employees}.get(eid)
        if employee is not None:
            self._audit("leave_clear_person", "清空个人请假", f"{employee.name}：清空全部请假")
            self.leave_message_box.appendPlainText(f"{employee.name}：清空全部请假")
        self.refresh_leave_list()

    def clear_all_leave_for_everyone(self):
        self.repo.clear_all_leaves()
        self._audit("leave_clear_all", "清空所有请假", "清空了所有人员请假记录")
        self.leave_message_box.appendPlainText("所有人员：清空全部请假")
        self.refresh_leave_list()

    def clear_leave_messages(self):
        self.leave_message_box.clear()

    def on_plan_calendar_clicked(self):
        self.current_plan_date = self.plan_calendar.selectedDate().toPython()
        self.load_plan_day_editor()

    def load_plan_day_editor(self):
        if self.current_plan_date is None:
            return
        d = self.current_plan_date
        tag = self.repo.load_day_tags().get(d, self._default_tag(d))
        manual = self.repo.load_manual_assignments("SL_MAIN")
        self.plan_date_label.setText(f"\u5f53\u524d\u65e5\u671f: {d.isoformat()}")
        default_label = "\u7279\u6b8a\u65e5\u671f" if self._default_tag(d) == "SPECIAL" else "\u5de5\u4f5c\u65e5"
        self.plan_default_label.setText(f"\u9ed8\u8ba4\u7c7b\u578b: {default_label}")
        self.plan_tag_combo.blockSignals(True)
        self.plan_tag_combo.setCurrentText("\u7279\u6b8a\u65e5\u671f" if tag == "SPECIAL" else "\u5de5\u4f5c\u65e5")
        self.plan_tag_combo.blockSignals(False)
        self.plan_leader_combo.blockSignals(True)
        idx = self.plan_leader_combo.findData(manual.get(d))
        self.plan_leader_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.plan_leader_combo.blockSignals(False)

    def save_selected_plan_tag(self):
        if self.current_plan_date is None:
            return
        d = self.current_plan_date
        tag = "SPECIAL" if self.plan_tag_combo.currentText() == "\u7279\u6b8a\u65e5\u671f" else "WORKDAY"
        self.repo.set_day_tag(d, tag)
        self._audit("plan_tag", "修改日期标签", f"{d.isoformat()}：{'特殊日期' if tag == 'SPECIAL' else '工作日'}")
        self.load_day_plan_table()

    def save_selected_plan_leader(self):
        if self.current_plan_date is None:
            return
        d = self.current_plan_date
        eid = self.plan_leader_combo.currentData()
        if eid is not None:
            self.repo.set_manual_assignment(d, "SL_MAIN", int(eid))
            employee = {e.id: e for e in self.employees}.get(int(eid))
            if employee is not None:
                self._audit("plan_leader", "指定特殊日主班", f"{d.isoformat()}：指定 {employee.name} 为双流主班")
        else:
            self.repo.clear_manual_assignment(d, "SL_MAIN")
            self._audit("plan_leader_clear", "清除特殊日主班", f"{d.isoformat()}：清除双流主班人工指定")
        self.load_day_plan_table()

    def clear_selected_plan_leader(self):
        start = self._month_start()
        end = self._month_end()
        self.repo.clear_manual_assignments_in_range("SL_MAIN", start, end)
        self.repo.clear_day_tags_in_range(start, end)
        self._audit("plan_reset", "重置当月计划日历", f"重置月份：{start.strftime('%Y-%m')} 的人工主班与日期标签")
        self.current_plan_date = start
        self.plan_calendar.setSelectedDate(QDate(start.year, start.month, start.day))
        self.plan_calendar.setCurrentPage(start.year, start.month)
        self.load_day_plan_table()

    def load_day_plan_table(self):
        start = self._month_start()
        end = self._month_end()
        tags = self.repo.load_day_tags()
        manual = self.repo.load_manual_assignments("SL_MAIN")
        days = []
        d = start
        while d <= end:
            days.append(d)
            d += timedelta(days=1)
        by_id = {e.id: e for e in self.employees}
        self.day_plan_table.setRowCount(len(days))
        manual_name_map: dict[date, str] = {}
        for i, d in enumerate(days):
            default_tag = self._default_tag(d)
            tag = tags.get(d, default_tag)
            leader_text = ""
            if d in manual and manual[d] in by_id:
                e = by_id[manual[d]]
                leader_text = f"{e.name}({e.team}-{e.squad})"
                manual_name_map[d] = e.name
            self.day_plan_table.setItem(i, 0, QTableWidgetItem(d.isoformat()))
            self.day_plan_table.setItem(i, 1, QTableWidgetItem("\u7279\u6b8a\u65e5\u671f" if default_tag == "SPECIAL" else "\u5de5\u4f5c\u65e5"))
            self.day_plan_table.setItem(i, 2, QTableWidgetItem("\u7279\u6b8a\u65e5\u671f" if tag == "SPECIAL" else "\u5de5\u4f5c\u65e5"))
            self.day_plan_table.setItem(i, 3, QTableWidgetItem(leader_text))
        if self.current_plan_date is None:
            self.current_plan_date = start
        elif self.current_plan_date < start or self.current_plan_date > end:
            self.current_plan_date = start
        self.plan_calendar.setSelectedDate(QDate(self.current_plan_date.year, self.current_plan_date.month, self.current_plan_date.day))
        self.plan_calendar.setCurrentPage(start.year, start.month)
        self.load_plan_day_editor()
        self._paint_plan_calendar()
        self.plan_calendar.set_manual_name_map(manual_name_map)
    def _collect_history_tail_from_xlsx(self, path: Path) -> list[dict]:
        wb = load_workbook(path, data_only=True)
        if "\u6392\u73ed\u7ed3\u679c" in wb.sheetnames:
            ws = wb["\u6392\u73ed\u7ed3\u679c"]
        elif "\u6392\u73ed\u8868" in wb.sheetnames:
            ws = wb["\u6392\u73ed\u8868"]
        else:
            ws = wb.active

        header_row = None
        idx: dict[str, int] = {}
        required = ["\u65e5\u671f", "\u53cc\u6d41\u4e3b\u73ed", "\u5929\u5e9c\u4e3b\u73ed", "\u673a\u961f\u603b\u8d1f\u8d23"]
        normalized_aliases = {
            "\u53cc\u6d41\u526f\u73ed(\u7a7a\u52e4)": "\u53cc\u6d41\u526f\u73ed\uff08\u7a7a\u52e4\uff09",
            "\u53cc\u6d41\u526f\u73ed(\u5730\u52e4)": "\u53cc\u6d41\u526f\u73ed\uff08\u5730\u52e4\uff09",
        }
        for r in range(1, min(ws.max_row, 10) + 1):
            header = []
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(r, c).value or "").strip()
                val = normalized_aliases.get(val, val)
                header.append(val)
            maybe_idx = {h: i + 1 for i, h in enumerate(header)}
            if all(h in maybe_idx for h in required):
                header_row = r
                idx = maybe_idx
                break

        if header_row is None:
            found_rows = []
            for r in range(1, min(ws.max_row, 10) + 1):
                found_rows.append([str(ws.cell(r, c).value or "").strip() for c in range(1, min(ws.max_column, 12) + 1)])
            raise ValueError(f"\u5bfc\u5165\u5386\u53f2\u6392\u73ed\u5931\u8d25\uff1a\u627e\u4e0d\u5230\u5fc5\u8981\u5217\u3002\u5de5\u4f5c\u8868={ws.title}\uff0c\u524d10\u884c={found_rows}")

        name_to_id = {e.name: e.id for e in self.employees}
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            d_raw = ws.cell(r, idx["\u65e5\u671f"]).value
            if d_raw is None:
                continue
            d_val = d_raw.date() if hasattr(d_raw, "date") else date.fromisoformat(str(d_raw)[:10])
            rows.append(
                (
                    d_val,
                    str(ws.cell(r, idx["\u53cc\u6d41\u4e3b\u73ed"]).value or "").strip(),
                    str(ws.cell(r, idx["\u5929\u5e9c\u4e3b\u73ed"]).value or "").strip(),
                    str(ws.cell(r, idx["\u673a\u961f\u603b\u8d1f\u8d23"]).value or "").strip(),
                )
            )
        rows.sort(key=lambda x: x[0])
        tail = rows[-3:]
        out = []
        for d_val, sl_name, tf_name, fleet_name in tail:
            if sl_name in name_to_id:
                out.append({"work_date": d_val.isoformat(), "position": "SL_MAIN", "employee_id": name_to_id[sl_name]})
            if tf_name in name_to_id:
                out.append({"work_date": d_val.isoformat(), "position": "TF_MAIN", "employee_id": name_to_id[tf_name]})
            if fleet_name in name_to_id:
                out.append({"work_date": d_val.isoformat(), "position": "FLEET_LEAD", "employee_id": name_to_id[fleet_name]})
        return out

    def _collect_year_stats_from_xlsx(self, path: Path) -> tuple[dict[int, int], dict[str, int]]:
        wb = load_workbook(path, data_only=True)
        person_year_totals: dict[int, int] = {}
        team_year_totals: dict[str, int] = {}
        employees_by_name = {e.name: e for e in self.employees}

        if "\u4eba\u5458\u5e74\u5ea6\u7edf\u8ba1" in wb.sheetnames:
            ws = wb["\u4eba\u5458\u5e74\u5ea6\u7edf\u8ba1"]
            for r in range(4, ws.max_row + 1):
                name = str(ws.cell(r, 1).value or "").strip()
                total = ws.cell(r, 4).value
                if name in employees_by_name and total not in (None, ""):
                    person_year_totals[employees_by_name[name].id] = int(total)

        if "\u5927\u961f\u5e74\u5ea6\u7edf\u8ba1" in wb.sheetnames:
            ws = wb["\u5927\u961f\u5e74\u5ea6\u7edf\u8ba1"]
            for r in range(4, ws.max_row + 1):
                team = str(ws.cell(r, 1).value or "").strip()
                total = ws.cell(r, 2).value
                if team and total not in (None, ""):
                    team_year_totals[team] = int(total)

        return person_year_totals, team_year_totals

    def _extract_schedule_month_key_from_xlsx(self, path: Path) -> str:
        wb = load_workbook(path, data_only=True)
        if "排班结果" in wb.sheetnames:
            ws = wb["排班结果"]
        elif "排班表" in wb.sheetnames:
            ws = wb["排班表"]
        else:
            ws = wb.active

        header_row = None
        date_col = None
        for r in range(1, min(ws.max_row, 10) + 1):
            header = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
            if "日期" in header:
                header_row = r
                date_col = header.index("日期") + 1
                break
        if header_row is None or date_col is None:
            raise ValueError("无法识别导入文件中的日期列")

        month_keys: set[str] = set()
        for r in range(header_row + 1, ws.max_row + 1):
            d_raw = ws.cell(r, date_col).value
            if d_raw in (None, ""):
                continue
            d_val = d_raw.date() if hasattr(d_raw, "date") else date.fromisoformat(str(d_raw)[:10])
            month_keys.add(f"{d_val.year:04d}-{d_val.month:02d}")
        if not month_keys:
            raise ValueError("导入文件中没有可识别的排班日期")
        if len(month_keys) != 1:
            raise ValueError(f"导入文件包含多个月份：{sorted(month_keys)}")
        return next(iter(month_keys))

    def import_history_from_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "\u5bfc\u5165\u5386\u53f2\u6392\u73ed", "", "Excel (*.xlsx *.xlsm)")
        if not file_path:
            return
        try:
            imported_month_key = self._extract_schedule_month_key_from_xlsx(Path(file_path))
            expected_month_key = self._previous_month_key()
            if imported_month_key != expected_month_key:
                raise ValueError(
                    f"\u5f53\u524d\u6392\u73ed\u6708\u4efd\u53ea\u80fd\u5f15\u7528\u4e0a\u4e2a\u6708\u7684\u5386\u53f2\u6392\u73ed\u3002"
                    f"\u5f53\u524d\u671f\u671b={expected_month_key}\uff0c\u4f46\u4f60\u9009\u62e9\u7684\u6587\u4ef6\u5c5e\u4e8e {imported_month_key}\u3002"
                )
            self.history_tail = self._collect_history_tail_from_xlsx(Path(file_path))
            self.history_person_year_totals, self.history_team_year_totals = self._collect_year_stats_from_xlsx(Path(file_path))
            self.history_source_locked = True
            self.history_source_type = "file"
            self.repo.save_history_import_cache("file", file_path, json.dumps(self.history_tail, ensure_ascii=False))
            self._audit("history_import", "导入上月历史排班", f"导入文件：{Path(file_path).name}，历史记录 {len(self.history_tail)} 条")
            QMessageBox.information(self, "\u6210\u529f", f"\u5df2\u5bfc\u5165\u5386\u53f2\u5c3e\u90e8\u8bb0\u5f55 {len(self.history_tail)} \u6761")
        except Exception as e:
            QMessageBox.critical(self, "\u5bfc\u5165\u5931\u8d25", str(e))

    def export_people_to_excel(self):
        default_name = f"people_{date.today().isoformat()}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "\u5bfc\u51fa\u4eba\u5458", default_name, "Excel (*.xlsx)")
        if not file_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "\u4eba\u5458\u4fe1\u606f"
            headers = ["\u5927\u961f", "\u4e2d\u961f", "\u59d3\u540d", "\u7c7b\u522b", "\u804c\u8d23"]
            ws.append(headers)
            for employee in self.repo.load_employees():
                ws.append([employee.team, employee.squad, employee.name, employee.duty_group, employee.role])
            widths = {"A": 10, "B": 12, "C": 16, "D": 10, "E": 18}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            wb.save(file_path)
            QMessageBox.information(self, "\u5bfc\u51fa\u5b8c\u6210", f"\u5df2\u5bfc\u51fa {len(self.repo.load_employees())} \u540d\u4eba\u5458\u5230\uff1a\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "\u5bfc\u51fa\u5931\u8d25", str(e))

    def import_db_month(self, month_key: str):
        file_path, _ = QFileDialog.getOpenFileName(self, f"\u5bfc\u5165 {month_key} \u5f53\u6708\u6570\u636e", "", "Excel (*.xlsx *.xlsm)")
        if not file_path:
            return
        try:
            self.repo.import_month_from_xlsx(month_key, Path(file_path), self.employees)
            self.refresh_db_management()
            self._audit("db_month_import", "导入数据库月份数据", f"导入月份：{month_key}，文件：{Path(file_path).name}")
            QMessageBox.information(self, "\u6210\u529f", f"{month_key} \u5f53\u6708\u6570\u636e\u5df2\u5bfc\u5165\u5e76\u4fdd\u5b58\u5230 {month_key}")
        except Exception as e:
            QMessageBox.critical(self, "\u5bfc\u5165\u5931\u8d25", str(e))

    def delete_db_month(self, month_key: str):
        year = self._month_start().year
        if month_key not in self.repo.list_saved_months(f"{year:04d}"):
            QMessageBox.information(self, "\u63d0\u793a", f"{month_key} \u76ee\u524d\u6ca1\u6709\u5df2\u4fdd\u5b58\u7684\u6570\u636e\u53ef\u5220\u9664")
            return
        reply = QMessageBox.question(
            self,
            "\u786e\u8ba4\u5220\u9664",
            f"\u786e\u5b9a\u5220\u9664 {month_key} \u7684\u6570\u636e\u5e93\u8bb0\u5f55\u5417\uff1f\u8fd9\u4f1a\u79fb\u9664\u8be5\u6708\u4fdd\u5b58\u7684\u6392\u73ed\u548c\u5e74\u5ea6\u5feb\u7167\u3002",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.repo.delete_month_data(month_key)
        self.refresh_db_management()
        self._audit("db_month_delete", "删除数据库月份数据", f"删除月份：{month_key}")
        QMessageBox.information(self, "\u6210\u529f", f"{month_key} \u6570\u636e\u5df2\u5220\u9664")

    def _auto_load_previous_month_history(self):
        prev_month_key = self._previous_month_key()
        tail = self.repo.latest_saved_tail(prev_month_key)
        if not tail:
            self.history_tail = []
            self.history_person_year_totals = {}
            self.history_team_year_totals = {}
            return
        self.history_tail = tail
        self.history_person_year_totals, self.history_team_year_totals = self.repo.load_month_year_snapshots(prev_month_key)

    def _previous_month_key(self) -> str:
        s = self._month_start()
        prev_end = s - timedelta(days=1)
        return f"{prev_end.year:04d}-{prev_end.month:02d}"

    def use_history_from_db(self):
        if self.history_source_type == "file":
            QMessageBox.information(self, "\u63d0\u793a", "\u5f53\u524d\u5df2\u5bfc\u5165\u5386\u53f2\u6392\u73ed\u6587\u4ef6\uff0c\u672c\u6b21\u6392\u73ed\u5c06\u4f18\u5148\u4f7f\u7528\u5bfc\u5165\u6587\u4ef6\u6570\u636e\u3002\u82e5\u8981\u6539\u4e3a\u6570\u636e\u5e93\u6570\u636e\uff0c\u8bf7\u5148\u5207\u6362\u6708\u4efd\u6216\u70b9\u51fb\u201c\u6e05\u9664\u5386\u53f2\u6570\u636e\u201d\u3002")
            return
        tail = self.repo.latest_saved_tail(self._previous_month_key())
        if not tail:
            QMessageBox.warning(self, "\u63d0\u793a", "\u6570\u636e\u5e93\u4e2d\u6ca1\u6709\u53ef\u5f15\u7528\u7684\u4e0a\u6708\u6392\u73ed\u8bb0\u5f55")
            return
        self.history_tail = tail
        self.history_person_year_totals, self.history_team_year_totals = self.repo.load_month_year_snapshots(
            self._previous_month_key()
        )
        self.history_source_locked = True
        self.history_source_type = "db"
        self.repo.save_history_import_cache("db", self._previous_month_key(), json.dumps(self.history_tail, ensure_ascii=False))
        self._audit("history_db_use", "引用排班数据库", f"引用上月数据：{self._previous_month_key()}，历史记录 {len(self.history_tail)} 条")
        QMessageBox.information(self, "\u6210\u529f", f"\u5df2\u4ece\u6570\u636e\u5e93\u5f15\u7528\u5386\u53f2\u5c3e\u90e8\u8bb0\u5f55 {len(self.history_tail)} \u6761")

    def clear_saved_history_data(self):
        reply = QMessageBox.question(
            self,
            "\u786e\u8ba4\u6e05\u9664",
            "\u8fd9\u4f1a\u6e05\u9664\u6570\u636e\u5e93\u4e2d\u7684\u6392\u73ed\u7ed3\u679c\u3001\u6708\u5ea6\u7edf\u8ba1\u3001\u5e74\u5ea6\u7edf\u8ba1\u548c\u5386\u53f2\u5f15\u7528\u7f13\u5b58\uff0c\u7528\u4e8e\u907f\u514d\u91cd\u590d\u53e0\u52a0\u3002\u662f\u5426\u7ee7\u7eed\uff1f",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.repo.clear_saved_schedule_stats()
        self.history_tail = []
        self.history_person_year_totals = {}
        self.history_team_year_totals = {}
        self.history_source_locked = False
        self.history_source_type = None
        self.render_summary()
        self.refresh_db_management()
        self._audit("history_clear", "清除历史数据", "清除了数据库中的排班结果、统计数据和历史引用缓存")
        QMessageBox.information(self, "\u6210\u529f", "\u5df2\u6e05\u9664\u6392\u73ed\u7edf\u8ba1\u5386\u53f2\u6570\u636e")

    def _run_schedule(self, use_result_overrides: bool, seed: int):
        if not self.history_source_locked:
            self._auto_load_previous_month_history()
        self.current_assignments = []
        self.current_logs = []
        self.log_box.clear()
        self.employees = self.repo.load_employees()
        overrides = self.repo.load_result_overrides() if use_result_overrides else {}
        result = SchedulerEngine().solve(
            self.employees,
            self._month_start(),
            self._month_end(),
            self.repo.load_leaves(),
            self.repo.load_day_tags(),
            self.repo.load_manual_assignments("SL_MAIN"),
            history_tail=self.history_tail,
            manual_overrides=overrides,
            rerun_seed=seed,
            person_year_baseline=self.history_person_year_totals,
            team_year_baseline=self.history_team_year_totals,
        )
        self.current_assignments = result.assignments
        self.current_logs = result.logs
        self.log_box.setPlainText("\n".join(result.logs) if result.logs else "排班完成")
        self.render_result_table()
        self.render_summary()
        self.result_status_label.setText(
            f"结果状态：已刷新（共 {len(self.current_assignments)} 条岗位安排，日志 {len(self.current_logs)} 条）"
        )
        # 每次重排后切换到“排班结果”页，确保界面立即显示最新结果。
        self.tabs.setCurrentWidget(self.tab_result)
        self.result_table.viewport().update()
        self.summary_table.viewport().update()

    def generate_schedule(self):
        self.rerun_seed = 0
        self._run_schedule(use_result_overrides=True, seed=self.rerun_seed)
        self._audit("schedule_generate", "自动排班", f"执行月份：{self._month_start().strftime('%Y-%m')}，生成 {len(self.current_assignments)} 条岗位安排")

    def regenerate_with_overrides(self):
        self.rerun_seed += 1
        self._run_schedule(use_result_overrides=True, seed=self.rerun_seed)
        self._audit("schedule_rerun_with_mod", "根据修改重新排班", f"执行月份：{self._month_start().strftime('%Y-%m')}，生成 {len(self.current_assignments)} 条岗位安排")

    def regenerate_plan_only(self):
        self.repo.clear_all_result_overrides()
        self.rerun_seed += 1
        self._run_schedule(use_result_overrides=False, seed=self.rerun_seed)
        self._audit("schedule_rerun_all", "全部重排", f"执行月份：{self._month_start().strftime('%Y-%m')}，生成 {len(self.current_assignments)} 条岗位安排")

    def _position_options(self, pos: str, d: date) -> list[tuple[str, int]]:
        leaves = self.repo.load_leaves()
        opts = []
        for e in self.employees:
            if not e.participate or (e.id, d) in leaves:
                continue
            if pos in ("SL_GROUND", "TF_GROUND") and e.duty_group != "\u5730\u52e4":
                continue
            if pos in ("SL_MAIN", "SL_AIR", "TF_MAIN") and e.duty_group != "\u7a7a\u52e4":
                continue
            opts.append((f"{e.name}({e.team}-{e.squad})", e.id))
        return opts

    def on_result_assignment_changed(self, d: date, pos: str, combo: QComboBox):
        eid = combo.currentData()
        if eid is None:
            self.repo.clear_result_override(d, pos)
            self._audit("result_override_clear", "清除排班结果人工修改", f"{d.isoformat()}：清除 {POSITION_LABELS[pos]} 的人工修改")
        else:
            self.repo.save_result_override(d, pos, int(eid))
            employee = {e.id: e for e in self.employees}.get(int(eid))
            if employee is not None:
                self._audit(
                    "result_override_set",
                    "修改排班结果",
                    f"{d.isoformat()}：{POSITION_LABELS[pos]} -> {employee.name}（{employee.team}-{employee.squad}）",
                )

    def render_result_table(self):
        by_id = {e.id: e for e in self.employees}
        rows_by_date = defaultdict(dict)
        for a in self.current_assignments:
            rows_by_date[a.work_date][a.position] = a
        all_dates = sorted(rows_by_date)
        overrides = self.repo.load_result_overrides()
        self.result_table.setRowCount(len(all_dates))
        for i, d in enumerate(all_dates):
            self.result_table.setItem(i, 0, QTableWidgetItem(d.isoformat()))
            self.result_table.setItem(i, 1, QTableWidgetItem(WEEKDAY_MAP[d.weekday()]))
            special = self.repo.load_day_tags().get(d, self._default_tag(d)) == "SPECIAL"
            self.result_table.setItem(i, 2, QTableWidgetItem("是" if special else "否"))
            for c, pos in RESULT_COL_POS.items():
                current = rows_by_date[d].get(pos)
                combo = QComboBox()
                combo.addItem("(空)", None)
                for label, eid in self._position_options(pos, d):
                    combo.addItem(label, eid)
                selected = overrides.get((d, pos), current.employee_id if current else None)
                idx = combo.findData(selected)
                combo.setCurrentIndex(idx if idx >= 0 else 0)
                combo.currentIndexChanged.connect(lambda _, dd=d, pp=pos, cb=combo: self.on_result_assignment_changed(dd, pp, cb))
                self.result_table.setCellWidget(i, c, combo)
            sl = rows_by_date[d].get("SL_MAIN")
            tf = rows_by_date[d].get("TF_MAIN")
            self.result_table.setItem(i, 9, QTableWidgetItem(by_id[sl.employee_id].team if sl and sl.employee_id in by_id else ""))
            self.result_table.setItem(i, 10, QTableWidgetItem(by_id[tf.employee_id].team if tf and tf.employee_id in by_id else ""))

    def save_current_schedule_data(self):
        if not self.current_assignments:
            QMessageBox.warning(self, "\u63d0\u793a", "\u8bf7\u5148\u751f\u6210\u6392\u73ed\u7ed3\u679c")
            return
        month_key = self._month_start().strftime("%Y-%m")
        self.repo.save_schedule_and_stats(self.current_assignments, self.employees, month_key)
        person_year_totals, team_year_totals = self._current_year_totals()
        self.repo.save_month_year_snapshots(month_key, person_year_totals, self.employees, team_year_totals)
        self._audit("schedule_save", "保存排班数据", f"保存月份：{month_key}，岗位安排 {len(self.current_assignments)} 条")
        QMessageBox.information(self, "\u6210\u529f", f"\u6392\u73ed\u7ed3\u679c\u4e0e\u7edf\u8ba1\u6570\u636e\u5df2\u4fdd\u5b58\u5230 {month_key}")
        self.render_summary()
        self.refresh_db_management()

    def render_summary(self):
        month_cnt = defaultdict(int)
        for a in self.current_assignments:
            month_cnt[a.employee_id] += 1
        person_year_totals, _team_year_totals = self._current_year_totals()
        year_cnt = person_year_totals
        self.summary_table.setRowCount(len(self.employees))
        for i, e in enumerate(self.employees):
            self.summary_table.setItem(i, 0, QTableWidgetItem(e.name))
            self.summary_table.setItem(i, 1, QTableWidgetItem(e.team))
            self.summary_table.setItem(i, 2, QTableWidgetItem(e.squad))
            self.summary_table.setItem(i, 3, QTableWidgetItem(str(month_cnt.get(e.id, 0))))
            self.summary_table.setItem(i, 4, QTableWidgetItem(str(year_cnt.get(e.id, 0))))

    def _current_year_totals(self) -> tuple[dict[int, int], dict[str, int]]:
        month_cnt = defaultdict(int)
        for a in self.current_assignments:
            month_cnt[a.employee_id] += 1
        year_key = self._month_start().strftime("%Y")
        db_person_totals = self.repo.yearly_totals(year_key)

        person_year_totals: dict[int, int] = {}
        all_person_ids = set(db_person_totals) | set(self.history_person_year_totals)
        for eid in all_person_ids:
            person_year_totals[eid] = max(
                db_person_totals.get(eid, 0),
                self.history_person_year_totals.get(eid, 0),
            )
        for eid, cnt in month_cnt.items():
            person_year_totals[eid] = person_year_totals.get(eid, 0) + cnt

        team_year_totals = defaultdict(int)
        by_id = {e.id: e for e in self.employees}
        for eid, cnt in person_year_totals.items():
            e = by_id.get(eid)
            if e is not None:
                team_year_totals[e.team] += cnt
        for team, cnt in self.history_team_year_totals.items():
            if cnt > team_year_totals.get(team, 0):
                team_year_totals[team] = cnt
        return person_year_totals, dict(team_year_totals)
    def on_export_csv(self):
        if not self.current_assignments:
            return
        path, _ = QFileDialog.getSaveFileName(self, "\u5bfc\u51faCSV", "schedule.csv", "CSV (*.csv)")
        if path:
            person_year_totals, team_year_totals = self._current_year_totals()
            export_csv(
                Path(path),
                self.current_assignments,
                self.employees,
                person_year_totals=person_year_totals,
                team_year_totals=team_year_totals,
            )

    def on_export_xlsx(self):
        if not self.current_assignments:
            return
        path, _ = QFileDialog.getSaveFileName(self, "\u5bfc\u51faExcel", "schedule.xlsx", "Excel (*.xlsx)")
        if path:
            person_year_totals, team_year_totals = self._current_year_totals()
            export_excel(
                Path(path),
                self.current_assignments,
                self.employees,
                self.repo.load_day_tags(),
                person_year_totals=person_year_totals,
                team_year_totals=team_year_totals,
            )


def run_app() -> None:
    app = QApplication(sys.argv)
    app.setApplicationName("Scheduler App")
    db_path = Path.home() / "AppData" / "Local" / "SchedulerApp" / "scheduler.db"
    repo = SchedulerRepository(db_path)
    repo.seed_if_empty()
    login = LoginDialog(repo, "登录排班系统")
    if login.exec() != QDialog.Accepted or login.account is None:
        sys.exit(0)
    repo.add_audit_log(login.account["id"], login.account["username"], "login_success", "登录成功", "成功登录排班系统")
    window = MainWindow(db_path, login.account)
    window.show()
    sys.exit(app.exec())
