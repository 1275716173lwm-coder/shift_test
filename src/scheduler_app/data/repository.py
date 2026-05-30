from __future__ import annotations

import hashlib
import sqlite3
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from contextlib import contextmanager
from openpyxl import load_workbook

from scheduler_app.core.models import Assignment, Employee


ROLE_MAP = {
    "leader": "大队长",
    "deputy_leader": "副大队长_空勤",
    "ground_admin": "中队书记",
    "member": "中队长",
    "squad_leader": "中队长",
    "squad_deputy": "副中队长",
}
GROUP_MAP = {"air": "空勤", "ground": "地勤"}
TEAM_MAP = {"A": "二", "B": "三", "C": "四"}
POSITION_NAME_TO_CODE = {
    "双流主班": "SL_MAIN",
    "双流副班（空勤）": "SL_AIR",
    "双流副班(空勤)": "SL_AIR",
    "双流副班（地勤）": "SL_GROUND",
    "双流副班(地勤)": "SL_GROUND",
    "天府主班": "TF_MAIN",
    "天府副班": "TF_GROUND",
    "天府副班(地勤)": "TF_GROUND",
    "机队总负责": "FLEET_LEAD",
}


def _count_duty_days_by_person(assignments: list[Assignment]) -> dict[int, int]:
    counts: dict[int, int] = defaultdict(int)
    seen: set[tuple[date, int]] = set()
    for a in assignments:
        key = (a.work_date, a.employee_id)
        if key in seen:
            continue
        seen.add(key)
        counts[a.employee_id] += 1
    return dict(counts)


def _count_duty_days_by_team(assignments: list[Assignment], employees: list[Employee]) -> dict[str, int]:
    by_id = {e.id: e for e in employees}
    counts: dict[str, int] = defaultdict(int)
    seen: set[tuple[date, int]] = set()
    for a in assignments:
        key = (a.work_date, a.employee_id)
        if key in seen:
            continue
        seen.add(key)
        employee = by_id.get(a.employee_id)
        if employee is not None:
            counts[employee.team] += 1
    return dict(counts)


class SchedulerRepository:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_schema()

    @contextmanager
    def _connect(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
            conn.commit()
        finally:
            conn.close()

    def _init_schema(self) -> None:
        with self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS employees(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    team TEXT NOT NULL,
                    squad TEXT NOT NULL DEFAULT '一中队',
                    duty_group TEXT NOT NULL,
                    role TEXT NOT NULL,
                    participate INTEGER NOT NULL DEFAULT 1,
                    order_index INTEGER NOT NULL DEFAULT 0
                );
                CREATE TABLE IF NOT EXISTS leaves(
                    employee_id INTEGER NOT NULL,
                    work_date TEXT NOT NULL,
                    PRIMARY KEY(employee_id, work_date)
                );
                CREATE TABLE IF NOT EXISTS day_tags(
                    work_date TEXT PRIMARY KEY,
                    tag TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS manual_assignments(
                    work_date TEXT NOT NULL,
                    position TEXT NOT NULL,
                    employee_id INTEGER NOT NULL,
                    PRIMARY KEY(work_date, position)
                );
                CREATE TABLE IF NOT EXISTS snapshots(
                    id TEXT PRIMARY KEY,
                    created_at TEXT NOT NULL,
                    note TEXT NOT NULL,
                    assignments_json TEXT NOT NULL,
                    logs_text TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS result_manual_overrides(
                    work_date TEXT NOT NULL,
                    position TEXT NOT NULL,
                    employee_id INTEGER NOT NULL,
                    created_at TEXT NOT NULL,
                    PRIMARY KEY(work_date, position)
                );
                CREATE TABLE IF NOT EXISTS saved_schedules(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    saved_at TEXT NOT NULL,
                    month_key TEXT NOT NULL,
                    work_date TEXT NOT NULL,
                    position TEXT NOT NULL,
                    employee_id INTEGER NOT NULL,
                    manual INTEGER NOT NULL DEFAULT 0
                );
                CREATE TABLE IF NOT EXISTS monthly_stats_history(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    saved_at TEXT NOT NULL,
                    month_key TEXT NOT NULL,
                    scope TEXT NOT NULL,
                    subject_id TEXT NOT NULL,
                    subject_name TEXT NOT NULL,
                    count_value INTEGER NOT NULL
                );
                CREATE TABLE IF NOT EXISTS yearly_stats_ledger(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    saved_at TEXT NOT NULL,
                    year_key TEXT NOT NULL,
                    employee_id INTEGER NOT NULL,
                    increment_value INTEGER NOT NULL
                );
                CREATE TABLE IF NOT EXISTS history_import_cache(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    imported_at TEXT NOT NULL,
                    source_type TEXT NOT NULL,
                    source_ref TEXT NOT NULL,
                    payload_json TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS month_year_snapshots(
                    month_key TEXT NOT NULL,
                    scope TEXT NOT NULL,
                    subject_id TEXT NOT NULL,
                    subject_name TEXT NOT NULL,
                    total_value INTEGER NOT NULL,
                    PRIMARY KEY(month_key, scope, subject_id)
                );
                CREATE TABLE IF NOT EXISTS accounts(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    is_admin INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS account_audit_logs(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    account_id INTEGER NOT NULL,
                    username_snapshot TEXT NOT NULL,
                    action_type TEXT NOT NULL,
                    action_label TEXT NOT NULL,
                    details_text TEXT NOT NULL,
                    created_at TEXT NOT NULL
                );
                """
            )
            cols = {r["name"] for r in conn.execute("PRAGMA table_info(employees)").fetchall()}
            if "squad" not in cols:
                conn.execute("ALTER TABLE employees ADD COLUMN squad TEXT NOT NULL DEFAULT '一中队'")
            if "order_index" not in cols:
                conn.execute("ALTER TABLE employees ADD COLUMN order_index INTEGER NOT NULL DEFAULT 0")

            rows = conn.execute("SELECT id, team, duty_group, role FROM employees").fetchall()
            for r in rows:
                team = TEAM_MAP.get(r["team"], r["team"])
                duty = GROUP_MAP.get(r["duty_group"], r["duty_group"])
                role = ROLE_MAP.get(r["role"], r["role"])
                conn.execute(
                    "UPDATE employees SET team=?, duty_group=?, role=? WHERE id=?",
                    (team, duty, role, r["id"]),
                )
            # Ensure every record has a stable order.
            ordered = conn.execute("SELECT id FROM employees ORDER BY order_index, id").fetchall()
            for idx, r in enumerate(ordered):
                conn.execute("UPDATE employees SET order_index=? WHERE id=?", (idx, r["id"]))

            admin_row = conn.execute("SELECT id FROM accounts WHERE username='admin111'").fetchone()
            if admin_row is None:
                now = datetime.now().isoformat(timespec="seconds")
                conn.execute(
                    "INSERT INTO accounts(username,password_hash,is_admin,created_at,updated_at) VALUES(?,?,?,?,?)",
                    ("admin111", self._hash_password("admin111"), 1, now, now),
                )

    @staticmethod
    def _hash_password(password: str) -> str:
        return hashlib.sha256(password.encode("utf-8")).hexdigest()

    def verify_login(self, username: str, password: str) -> dict | None:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT id, username, is_admin, password_hash FROM accounts WHERE username=?",
                (username.strip(),),
            ).fetchone()
            if row is None or row["password_hash"] != self._hash_password(password):
                return None
            return {"id": row["id"], "username": row["username"], "is_admin": bool(row["is_admin"])}

    def verify_admin_credentials(self, username: str, password: str) -> bool:
        account = self.verify_login(username, password)
        return bool(account and account["is_admin"])

    def list_accounts(self) -> list[dict]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT id, username, is_admin, created_at, updated_at FROM accounts ORDER BY is_admin DESC, username COLLATE NOCASE"
            ).fetchall()
            return [
                {
                    "id": row["id"],
                    "username": row["username"],
                    "is_admin": bool(row["is_admin"]),
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                }
                for row in rows
            ]

    def create_account(self, username: str, password: str, is_admin: bool) -> None:
        username = username.strip()
        if not username:
            raise ValueError("用户名不能为空")
        if not password:
            raise ValueError("密码不能为空")
        with self._connect() as conn:
            exists = conn.execute("SELECT 1 FROM accounts WHERE username=?", (username,)).fetchone()
            if exists is not None:
                raise ValueError("用户名已存在")
            now = datetime.now().isoformat(timespec="seconds")
            conn.execute(
                "INSERT INTO accounts(username,password_hash,is_admin,created_at,updated_at) VALUES(?,?,?,?,?)",
                (username, self._hash_password(password), 1 if is_admin else 0, now, now),
            )

    def update_account(self, account_id: int, username: str, password: str | None, is_admin: bool) -> None:
        username = username.strip()
        if not username:
            raise ValueError("用户名不能为空")
        with self._connect() as conn:
            row = conn.execute("SELECT id, is_admin FROM accounts WHERE id=?", (account_id,)).fetchone()
            if row is None:
                raise ValueError("账户不存在")
            exists = conn.execute("SELECT 1 FROM accounts WHERE username=? AND id<>?", (username, account_id)).fetchone()
            if exists is not None:
                raise ValueError("用户名已存在")
            if row["is_admin"] and not is_admin:
                admin_count = conn.execute("SELECT COUNT(*) AS n FROM accounts WHERE is_admin=1").fetchone()["n"]
                if admin_count <= 1:
                    raise ValueError("至少需要保留一个管理员账户")
            now = datetime.now().isoformat(timespec="seconds")
            if password:
                conn.execute(
                    "UPDATE accounts SET username=?, password_hash=?, is_admin=?, updated_at=? WHERE id=?",
                    (username, self._hash_password(password), 1 if is_admin else 0, now, account_id),
                )
            else:
                conn.execute(
                    "UPDATE accounts SET username=?, is_admin=?, updated_at=? WHERE id=?",
                    (username, 1 if is_admin else 0, now, account_id),
                )

    def delete_account(self, account_id: int) -> None:
        with self._connect() as conn:
            row = conn.execute("SELECT id, is_admin FROM accounts WHERE id=?", (account_id,)).fetchone()
            if row is None:
                raise ValueError("账户不存在")
            if row["is_admin"]:
                admin_count = conn.execute("SELECT COUNT(*) AS n FROM accounts WHERE is_admin=1").fetchone()["n"]
                if admin_count <= 1:
                    raise ValueError("至少需要保留一个管理员账户")
            conn.execute("DELETE FROM accounts WHERE id=?", (account_id,))

    def add_audit_log(self, account_id: int, username_snapshot: str, action_type: str, action_label: str, details_text: str) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO account_audit_logs(account_id,username_snapshot,action_type,action_label,details_text,created_at)
                VALUES(?,?,?,?,?,?)
                """,
                (
                    account_id,
                    username_snapshot,
                    action_type,
                    action_label,
                    details_text,
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )

    def load_audit_logs(self, account_id: int | None = None) -> list[dict]:
        with self._connect() as conn:
            if account_id is None:
                rows = conn.execute(
                    """
                    SELECT id, account_id, username_snapshot, action_type, action_label, details_text, created_at
                    FROM account_audit_logs
                    ORDER BY created_at DESC, id DESC
                    """
                ).fetchall()
            else:
                rows = conn.execute(
                    """
                    SELECT id, account_id, username_snapshot, action_type, action_label, details_text, created_at
                    FROM account_audit_logs
                    WHERE account_id=?
                    ORDER BY created_at DESC, id DESC
                    """,
                    (account_id,),
                ).fetchall()
            return [
                {
                    "id": row["id"],
                    "account_id": row["account_id"],
                    "username_snapshot": row["username_snapshot"],
                    "action_type": row["action_type"],
                    "action_label": row["action_label"],
                    "details_text": row["details_text"],
                    "created_at": row["created_at"],
                }
                for row in rows
            ]

    def list_audit_accounts(self) -> list[dict]:
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT account_id, username_snapshot, MAX(created_at) AS last_at
                FROM account_audit_logs
                GROUP BY account_id, username_snapshot
                ORDER BY last_at DESC, username_snapshot COLLATE NOCASE
                """
            ).fetchall()
            return [
                {
                    "account_id": row["account_id"],
                    "username_snapshot": row["username_snapshot"],
                }
                for row in rows
            ]

    def seed_if_empty(self) -> None:
        with self._connect() as conn:
            n = conn.execute("SELECT COUNT(*) n FROM employees").fetchone()["n"]
            if n:
                return
            seed = [
                ("二大队长", "二", "一中队", "空勤", "大队长", 1),
                ("二副大队长空", "二", "一中队", "空勤", "副大队长_空勤", 1),
                ("二副大队长地", "二", "二中队", "地勤", "副大队长_地勤", 1),
                ("二党总支书记", "二", "二中队", "地勤", "党总支书记", 1),
                ("二中队长", "二", "三中队", "空勤", "中队长", 1),
                ("二副中队长", "二", "三中队", "空勤", "副中队长", 1),
                ("二中队书记", "二", "四中队", "地勤", "中队书记", 1),
                ("三大队长", "三", "一中队", "空勤", "大队长", 1),
                ("三副大队长空", "三", "一中队", "空勤", "副大队长_空勤", 1),
                ("三副大队长地", "三", "二中队", "地勤", "副大队长_地勤", 1),
                ("三党总支书记", "三", "二中队", "地勤", "党总支书记", 1),
                ("三中队长", "三", "三中队", "空勤", "中队长", 1),
                ("三副中队长", "三", "三中队", "空勤", "副中队长", 1),
                ("三中队书记", "三", "四中队", "地勤", "中队书记", 1),
                ("四大队长", "四", "一中队", "空勤", "大队长", 1),
                ("四副大队长空", "四", "一中队", "空勤", "副大队长_空勤", 1),
                ("四副大队长地", "四", "二中队", "地勤", "副大队长_地勤", 1),
                ("四党总支书记", "四", "二中队", "地勤", "党总支书记", 1),
                ("四中队长", "四", "三中队", "空勤", "中队长", 1),
                ("四副中队长", "四", "三中队", "空勤", "副中队长", 1),
                ("四中队书记", "四", "四中队", "地勤", "中队书记", 1),
            ]
            for idx, row in enumerate(seed):
                conn.execute(
                    "INSERT INTO employees(name,team,squad,duty_group,role,participate,order_index) VALUES(?,?,?,?,?,?,?)",
                    (*row, idx),
                )

    def load_employees(self) -> list[Employee]:
        with self._connect() as conn:
            rows = conn.execute("SELECT * FROM employees ORDER BY order_index, id").fetchall()
            return [Employee(id=r["id"], name=r["name"], team=r["team"], squad=r["squad"], duty_group=r["duty_group"], role=r["role"], participate=bool(r["participate"])) for r in rows]

    def add_employee(self, name: str, team: str, squad: str, duty_group: str, role: str, participate: bool) -> None:
        with self._connect() as conn:
            next_order = conn.execute("SELECT COALESCE(MAX(order_index), -1) + 1 AS n FROM employees").fetchone()["n"]
            conn.execute(
                "INSERT INTO employees(name,team,squad,duty_group,role,participate,order_index) VALUES(?,?,?,?,?,?,?)",
                (name, team, squad, duty_group, role, 1 if participate else 0, next_order),
            )

    def replace_employees_from_xlsx(self, xlsx_path: Path) -> int:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        def norm_text(val) -> str:
            return (
                str(val or "")
                .replace("\n", "")
                .replace("\r", "")
                .replace(" ", "")
                .replace("\u3000", "")
                .strip()
            )

        def canonical_header(val) -> str:
            raw = norm_text(val)
            alias_groups = {
                "姓名": {"姓名", "人员姓名", "名字", "人员"},
                "大队": {"大队", "所属大队", "大队别"},
                "中队": {"中队", "所属中队", "中队别", "分队", "所属分队"},
                "类别": {"类别", "组别", "岗位类别"},
                "职责": {"职责", "职务", "岗位", "岗位职责"},
                "参与排班": {"参与排班", "是否参与排班", "参与", "排班"},
            }
            for canonical, aliases in alias_groups.items():
                if raw in aliases:
                    return canonical
            return raw

        def norm_team(val: str) -> str:
            raw = norm_text(val)
            alias_map = {
                "二": "二",
                "二大队": "二",
                "2": "二",
                "三": "三",
                "三大队": "三",
                "3": "三",
                "四": "四",
                "四大队": "四",
                "4": "四",
            }
            return alias_map.get(raw, raw)

        def norm_squad(val: str) -> str:
            raw = norm_text(val)
            alias_map = {
                "直属": "直属",
                "直属队": "直属",
                "直": "直属",
                "一": "一中队",
                "一中队": "一中队",
                "1队": "一中队",
                "1中队": "一中队",
                "二": "二中队",
                "二中队": "二中队",
                "2队": "二中队",
                "2中队": "二中队",
                "三": "三中队",
                "三中队": "三中队",
                "3队": "三中队",
                "3中队": "三中队",
                "四": "四中队",
                "四中队": "四中队",
                "4队": "四中队",
                "4中队": "四中队",
            }
            return alias_map.get(raw, raw)

        required = ["姓名", "大队", "中队", "类别", "职责"]
        header_row = None
        idx: dict[str, int] = {}
        for row_no in range(1, min(ws.max_row, 10) + 1):
            headers = [canonical_header(ws.cell(row_no, c).value) for c in range(1, ws.max_column + 1)]
            candidate_idx = {h: i + 1 for i, h in enumerate(headers) if h}
            if all(col in candidate_idx for col in required):
                header_row = row_no
                idx = candidate_idx
                break

        if header_row is None:
            preview_rows = []
            for row_no in range(1, min(ws.max_row, 5) + 1):
                preview_rows.append(
                    [norm_text(ws.cell(row_no, c).value) for c in range(1, min(ws.max_column, 8) + 1)]
                )
            raise ValueError(f"Excel缺少必要列: {', '.join(required)}；前几行内容={preview_rows}")

        missing = [c for c in required if c not in idx]
        if missing:
            raise ValueError(f"Excel缺少必要列: {', '.join(missing)}")

        participate_col = idx.get("参与排班")

        rows_to_insert: list[tuple[str, str, str, str, str, int, int]] = []
        order_index = 0
        for r in range(header_row + 1, ws.max_row + 1):
            name = norm_text(ws.cell(r, idx["姓名"]).value)
            team = norm_team(ws.cell(r, idx["大队"]).value)
            squad = norm_squad(ws.cell(r, idx["中队"]).value)
            duty_group = norm_text(ws.cell(r, idx["类别"]).value)
            role = norm_text(ws.cell(r, idx["职责"]).value)
            if not name:
                continue
            if not team:
                team = "二"
            if not squad:
                squad = "直属"
            if not duty_group:
                duty_group = "空勤"
            if not role:
                role = "中队长"

            participate = 1
            if participate_col:
                raw = str(ws.cell(r, participate_col).value or "").strip()
                if raw in {"0", "否", "false", "False", "不参与"}:
                    participate = 0

            rows_to_insert.append((name, team, squad, duty_group, role, participate, order_index))
            order_index += 1

        with self._connect() as conn:
            conn.execute("DELETE FROM manual_assignments")
            conn.execute("DELETE FROM leaves")
            conn.execute("DELETE FROM employees")
            conn.execute("DELETE FROM sqlite_sequence WHERE name='employees'")
            for row in rows_to_insert:
                conn.execute(
                    "INSERT INTO employees(name,team,squad,duty_group,role,participate,order_index) VALUES(?,?,?,?,?,?,?)",
                    row,
                )
        return len(rows_to_insert)

    def update_employee_meta(self, employee_id: int, team: str, squad: str, duty_group: str, role: str) -> None:
        with self._connect() as conn:
            conn.execute(
                "UPDATE employees SET team=?, squad=?, duty_group=?, role=? WHERE id=?",
                (team, squad, duty_group, role, employee_id),
            )

    def delete_employee(self, employee_id: int) -> None:
        with self._connect() as conn:
            conn.execute("DELETE FROM leaves WHERE employee_id=?", (employee_id,))
            conn.execute("DELETE FROM employees WHERE id=?", (employee_id,))

    def set_participate(self, employee_id: int, participate: bool) -> None:
        with self._connect() as conn:
            conn.execute("UPDATE employees SET participate=? WHERE id=?", (1 if participate else 0, employee_id))

    def reorder_employees(self, employee_ids_in_order: list[int]) -> None:
        with self._connect() as conn:
            for idx, eid in enumerate(employee_ids_in_order):
                conn.execute("UPDATE employees SET order_index=? WHERE id=?", (idx, eid))

    def load_leaves(self) -> set[tuple[int, date]]:
        with self._connect() as conn:
            return {(r["employee_id"], date.fromisoformat(r["work_date"])) for r in conn.execute("SELECT * FROM leaves").fetchall()}

    def toggle_leave(self, employee_id: int, d: date) -> bool:
        with self._connect() as conn:
            row = conn.execute("SELECT 1 FROM leaves WHERE employee_id=? AND work_date=?", (employee_id, d.isoformat())).fetchone()
            if row:
                conn.execute("DELETE FROM leaves WHERE employee_id=? AND work_date=?", (employee_id, d.isoformat())); return False
            conn.execute("INSERT INTO leaves(employee_id,work_date) VALUES(?,?)", (employee_id, d.isoformat())); return True

    def remove_leave(self, employee_id: int, d: date) -> None:
        with self._connect() as conn:
            conn.execute("DELETE FROM leaves WHERE employee_id=? AND work_date=?", (employee_id, d.isoformat()))

    def clear_leaves_for_employee(self, employee_id: int) -> None:
        with self._connect() as conn:
            conn.execute("DELETE FROM leaves WHERE employee_id=?", (employee_id,))

    def clear_all_leaves(self) -> None:
        with self._connect() as conn:
            conn.execute("DELETE FROM leaves")

    def set_day_tag(self, d: date, tag: str) -> None:
        with self._connect() as conn:
            conn.execute("INSERT OR REPLACE INTO day_tags(work_date,tag) VALUES(?,?)", (d.isoformat(), tag))

    def load_day_tags(self) -> dict[date, str]:
        with self._connect() as conn:
            return {date.fromisoformat(r["work_date"]): r["tag"] for r in conn.execute("SELECT * FROM day_tags").fetchall()}

    def clear_day_tags_in_range(self, start: date, end: date) -> None:
        with self._connect() as conn:
            conn.execute(
                "DELETE FROM day_tags WHERE work_date>=? AND work_date<=?",
                (start.isoformat(), end.isoformat()),
            )

    def set_manual_assignment(self, d: date, position: str, employee_id: int) -> None:
        with self._connect() as conn:
            conn.execute("INSERT OR REPLACE INTO manual_assignments(work_date,position,employee_id) VALUES(?,?,?)", (d.isoformat(), position, employee_id))

    def clear_manual_assignment(self, d: date, position: str) -> None:
        with self._connect() as conn:
            conn.execute(
                "DELETE FROM manual_assignments WHERE work_date=? AND position=?",
                (d.isoformat(), position),
            )

    def clear_all_manual_assignments(self, position: str) -> None:
        with self._connect() as conn:
            conn.execute(
                "DELETE FROM manual_assignments WHERE position=?",
                (position,),
            )

    def clear_manual_assignments_in_range(self, position: str, start: date, end: date) -> None:
        with self._connect() as conn:
            conn.execute(
                "DELETE FROM manual_assignments WHERE position=? AND work_date>=? AND work_date<=?",
                (position, start.isoformat(), end.isoformat()),
            )

    def load_manual_assignments(self, position: str) -> dict[date, int]:
        with self._connect() as conn:
            return {date.fromisoformat(r["work_date"]): r["employee_id"] for r in conn.execute("SELECT * FROM manual_assignments WHERE position=?", (position,)).fetchall()}

    def save_result_override(self, d: date, position: str, employee_id: int) -> None:
        with self._connect() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO result_manual_overrides(work_date,position,employee_id,created_at) VALUES(?,?,?,?)",
                (d.isoformat(), position, employee_id, datetime.now().isoformat(timespec="seconds")),
            )

    def clear_result_override(self, d: date, position: str) -> None:
        with self._connect() as conn:
            conn.execute(
                "DELETE FROM result_manual_overrides WHERE work_date=? AND position=?",
                (d.isoformat(), position),
            )

    def clear_all_result_overrides(self) -> None:
        with self._connect() as conn:
            conn.execute("DELETE FROM result_manual_overrides")

    def load_result_overrides(self) -> dict[tuple[date, str], int]:
        with self._connect() as conn:
            rows = conn.execute("SELECT work_date, position, employee_id FROM result_manual_overrides").fetchall()
            return {(date.fromisoformat(r["work_date"]), r["position"]): r["employee_id"] for r in rows}

    def save_schedule_and_stats(self, assignments: list[Assignment], employees: list[Employee], month_key: str) -> None:
        by_id = {e.id: e for e in employees}
        saved_at = datetime.now().isoformat(timespec="seconds")
        month_person = _count_duty_days_by_person(assignments)
        month_team = _count_duty_days_by_team(assignments, employees)
        year_key = month_key[:4]
        with self._connect() as conn:
            for a in assignments:
                conn.execute(
                    "INSERT INTO saved_schedules(saved_at,month_key,work_date,position,employee_id,manual) VALUES(?,?,?,?,?,?)",
                    (saved_at, month_key, a.work_date.isoformat(), a.position, a.employee_id, 1 if a.manual else 0),
                )
            for eid, cnt in month_person.items():
                e = by_id.get(eid)
                if not e:
                    continue
                conn.execute(
                    "INSERT INTO monthly_stats_history(saved_at,month_key,scope,subject_id,subject_name,count_value) VALUES(?,?,?,?,?,?)",
                    (saved_at, month_key, "person", str(eid), e.name, cnt),
                )
                conn.execute(
                    "INSERT INTO yearly_stats_ledger(saved_at,year_key,employee_id,increment_value) VALUES(?,?,?,?)",
                    (saved_at, year_key, eid, cnt),
                )
            for team, cnt in month_team.items():
                conn.execute(
                    "INSERT INTO monthly_stats_history(saved_at,month_key,scope,subject_id,subject_name,count_value) VALUES(?,?,?,?,?,?)",
                    (saved_at, month_key, "team", team, team, cnt),
                )

    def save_month_year_snapshots(self, month_key: str, person_year_totals: dict[int, int], employees: list[Employee], team_year_totals: dict[str, int] | None = None) -> None:
        by_id = {e.id: e for e in employees}
        computed_team_totals = defaultdict(int)
        for eid, total in person_year_totals.items():
            e = by_id.get(eid)
            if e is not None:
                computed_team_totals[e.team] += total
        final_team_totals = dict(computed_team_totals)
        if team_year_totals:
            for team, total in team_year_totals.items():
                if total > final_team_totals.get(team, 0):
                    final_team_totals[team] = total
        with self._connect() as conn:
            conn.execute("DELETE FROM month_year_snapshots WHERE month_key=?", (month_key,))
            for eid, total in person_year_totals.items():
                e = by_id.get(eid)
                if e is None:
                    continue
                conn.execute(
                    "INSERT INTO month_year_snapshots(month_key,scope,subject_id,subject_name,total_value) VALUES(?,?,?,?,?)",
                    (month_key, "person", str(eid), e.name, int(total)),
                )
            for team, total in final_team_totals.items():
                conn.execute(
                    "INSERT INTO month_year_snapshots(month_key,scope,subject_id,subject_name,total_value) VALUES(?,?,?,?,?)",
                    (month_key, "team", team, team, int(total)),
                )

    def load_month_year_snapshots(self, month_key: str) -> tuple[dict[int, int], dict[str, int]]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT scope, subject_id, total_value FROM month_year_snapshots WHERE month_key=?",
                (month_key,),
            ).fetchall()
        person_totals: dict[int, int] = {}
        team_totals: dict[str, int] = {}
        for r in rows:
            if r["scope"] == "person":
                person_totals[int(r["subject_id"])] = int(r["total_value"])
            elif r["scope"] == "team":
                team_totals[str(r["subject_id"])] = int(r["total_value"])
        return person_totals, team_totals

    def yearly_totals(self, year_key: str) -> dict[int, int]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT employee_id, COALESCE(SUM(increment_value),0) AS c FROM yearly_stats_ledger WHERE year_key=? GROUP BY employee_id",
                (year_key,),
            ).fetchall()
            return {int(r["employee_id"]): int(r["c"]) for r in rows}

    def monthly_person_totals(self, month_key: str) -> dict[int, int]:
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT subject_id, count_value
                FROM monthly_stats_history
                WHERE month_key=? AND scope='person'
                """,
                (month_key,),
            ).fetchall()
        return {int(r["subject_id"]): int(r["count_value"]) for r in rows}

    def monthly_team_totals(self, month_key: str) -> dict[str, int]:
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT subject_id, count_value
                FROM monthly_stats_history
                WHERE month_key=? AND scope='team'
                """,
                (month_key,),
            ).fetchall()
        return {str(r["subject_id"]): int(r["count_value"]) for r in rows}

    def clear_saved_schedule_stats(self) -> None:
        with self._connect() as conn:
            conn.execute("DELETE FROM saved_schedules")
            conn.execute("DELETE FROM monthly_stats_history")
            conn.execute("DELETE FROM yearly_stats_ledger")
            conn.execute("DELETE FROM history_import_cache")
            conn.execute("DELETE FROM month_year_snapshots")

    def list_saved_months(self, year_key: str) -> set[str]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT DISTINCT month_key FROM saved_schedules WHERE month_key LIKE ?",
                (f"{year_key}-%",),
            ).fetchall()
        return {str(r["month_key"]) for r in rows}

    def delete_month_data(self, month_key: str) -> None:
        with self._connect() as conn:
            saved_at_rows = conn.execute(
                "SELECT DISTINCT saved_at FROM monthly_stats_history WHERE month_key=? AND scope='person'",
                (month_key,),
            ).fetchall()
            saved_ats = [str(r["saved_at"]) for r in saved_at_rows]
            conn.execute("DELETE FROM saved_schedules WHERE month_key=?", (month_key,))
            conn.execute("DELETE FROM monthly_stats_history WHERE month_key=?", (month_key,))
            conn.execute("DELETE FROM month_year_snapshots WHERE month_key=?", (month_key,))
            for saved_at in saved_ats:
                conn.execute("DELETE FROM yearly_stats_ledger WHERE saved_at=?", (saved_at,))

    def import_month_from_xlsx(self, month_key: str, xlsx_path: Path, employees: list[Employee]) -> None:
        wb = load_workbook(xlsx_path, data_only=True)
        if "排班结果" in wb.sheetnames:
            ws = wb["排班结果"]
        elif "排班表" in wb.sheetnames:
            ws = wb["排班表"]
        else:
            ws = wb.active

        header_row = None
        idx: dict[str, int] = {}
        required = ["日期", "双流主班", "天府主班", "机队总负责"]
        for r in range(1, min(ws.max_row, 10) + 1):
            header = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
            maybe_idx = {h: i + 1 for i, h in enumerate(header)}
            if all(h in maybe_idx for h in required):
                header_row = r
                idx = maybe_idx
                break
        if header_row is None:
            raise ValueError("导入月份数据失败：找不到排班结果表头")

        name_to_id = {e.name: e.id for e in employees}
        assignments: list[Assignment] = []
        imported_month_keys: set[str] = set()
        for r in range(header_row + 1, ws.max_row + 1):
            d_raw = ws.cell(r, idx["日期"]).value
            if d_raw in (None, ""):
                continue
            d_val = d_raw.date() if hasattr(d_raw, "date") else date.fromisoformat(str(d_raw)[:10])
            imported_month_keys.add(f"{d_val.year:04d}-{d_val.month:02d}")
            for col_name, pos_code in POSITION_NAME_TO_CODE.items():
                col_idx = idx.get(col_name)
                if not col_idx:
                    continue
                person_name = str(ws.cell(r, col_idx).value or "").strip()
                if person_name and person_name in name_to_id:
                    assignments.append(Assignment(d_val, pos_code, name_to_id[person_name], manual=False))
        if not imported_month_keys:
            raise ValueError("导入月份数据失败：文件中没有可识别的排班日期")
        if len(imported_month_keys) != 1:
            raise ValueError(f"导入月份数据失败：文件包含多个月份 {sorted(imported_month_keys)}")
        imported_month_key = next(iter(imported_month_keys))
        if imported_month_key != month_key:
            raise ValueError(
                f"导入月份不匹配：你当前导入到 {month_key}，但文件实际是 {imported_month_key}。"
            )

        person_year_totals: dict[int, int] = {}
        team_year_totals: dict[str, int] = {}
        by_name = {e.name: e for e in employees}
        if "人员年度统计" in wb.sheetnames:
            ws = wb["人员年度统计"]
            for r in range(4, ws.max_row + 1):
                name = str(ws.cell(r, 1).value or "").strip()
                total = ws.cell(r, 4).value
                if name in by_name and total not in (None, ""):
                    person_year_totals[by_name[name].id] = int(total)
        if "大队年度统计" in wb.sheetnames:
            ws = wb["大队年度统计"]
            for r in range(4, ws.max_row + 1):
                team = str(ws.cell(r, 1).value or "").strip()
                total = ws.cell(r, 2).value
                if team and total not in (None, ""):
                    team_year_totals[team] = int(total)

        self.delete_month_data(month_key)
        self.save_schedule_and_stats(assignments, employees, month_key)
        self.save_month_year_snapshots(month_key, person_year_totals, employees, team_year_totals)

    def save_history_import_cache(self, source_type: str, source_ref: str, payload_json: str) -> None:
        with self._connect() as conn:
            conn.execute(
                "INSERT INTO history_import_cache(imported_at,source_type,source_ref,payload_json) VALUES(?,?,?,?)",
                (datetime.now().isoformat(timespec="seconds"), source_type, source_ref, payload_json),
            )

    def latest_history_import_cache(self) -> dict | None:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT imported_at,source_type,source_ref,payload_json FROM history_import_cache ORDER BY id DESC LIMIT 1"
            ).fetchone()
            return dict(row) if row else None

    def latest_saved_tail(self, month_key: str) -> list[dict]:
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT work_date, position, employee_id
                FROM saved_schedules
                WHERE month_key=?
                ORDER BY id DESC
                """,
                (month_key,),
            ).fetchall()
        if not rows:
            return []
        latest_by_day_pos = {}
        for r in rows:
            k = (r["work_date"], r["position"])
            if k not in latest_by_day_pos:
                latest_by_day_pos[k] = dict(r)
        all_dates = sorted({k[0] for k in latest_by_day_pos})
        tail = all_dates[-3:]
        out = []
        for d in tail:
            for pos in ("SL_MAIN", "TF_MAIN", "FLEET_LEAD"):
                row = latest_by_day_pos.get((d, pos))
                if row:
                    out.append(row)
        return out

    def save_snapshot(self, snapshot_id: str, note: str, assignments: list[Assignment], logs: list[str]) -> None:
        import json
        payload = [{"work_date": a.work_date.isoformat(), "position": a.position, "employee_id": a.employee_id, "manual": a.manual} for a in assignments]
        with self._connect() as conn:
            conn.execute("INSERT INTO snapshots(id,created_at,note,assignments_json,logs_text) VALUES(?,?,?,?,?)", (snapshot_id, datetime.now().isoformat(timespec="seconds"), note, json.dumps(payload, ensure_ascii=False), "\n".join(logs)))

    def list_snapshots(self) -> list[dict]:
        with self._connect() as conn:
            return [dict(r) for r in conn.execute("SELECT id,created_at,note FROM snapshots ORDER BY created_at DESC").fetchall()]
