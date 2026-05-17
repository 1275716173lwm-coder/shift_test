from __future__ import annotations

import csv
from collections import defaultdict
from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from scheduler_app.core.models import Assignment, Employee, POSITION_LABELS


POSITION_EXPORT_ORDER = [
    "SL_MAIN",
    "SL_AIR",
    "SL_GROUND",
    "TF_MAIN",
    "TF_GROUND",
    "FLEET_LEAD",
]


def _person_position_dates(assignments: list[Assignment]) -> dict[int, dict[str, list[str]]]:
    person_dates: dict[int, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for a in sorted(assignments, key=lambda x: (x.employee_id, x.position, x.work_date)):
        person_dates[a.employee_id][a.position].append(f"{a.work_date.month}-{a.work_date.day}")
    return person_dates


def export_csv(
    path: Path,
    assignments: list[Assignment],
    employees: list[Employee],
    person_year_totals: dict[int, int] | None = None,
    team_year_totals: dict[str, int] | None = None,
) -> None:
    by_id = {e.id: e for e in employees}
    person_year_totals = person_year_totals or {}
    team_year_totals = team_year_totals or {}
    person_month_totals = defaultdict(int)
    team_month_totals = defaultdict(int)
    for a in assignments:
        person_month_totals[a.employee_id] += 1
        team_month_totals[by_id[a.employee_id].team] += 1
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        person_dates = _person_position_dates(assignments)
        w.writerow(["日期", "岗位", "姓名", "大队", "中队", "类别", "职责", "人工"])
        for a in sorted(assignments, key=lambda x: (x.work_date, x.position)):
            e = by_id[a.employee_id]
            w.writerow([
                a.work_date.isoformat(),
                POSITION_LABELS.get(a.position, a.position),
                e.name,
                e.team,
                e.squad,
                e.duty_group,
                e.role,
                int(a.manual),
            ])
        w.writerow([])
        w.writerow(["人员统计"])
        w.writerow(["姓名", "大队", "中队", "月度值班总数", "年度值班总数"])
        for e in employees:
            w.writerow([e.name, e.team, e.squad, person_month_totals[e.id], person_year_totals.get(e.id, 0)])
        w.writerow([])
        w.writerow(["大队统计"])
        w.writerow(["大队", "月度值班总数", "年度值班总数"])
        all_teams = sorted({e.team for e in employees})
        for team in all_teams:
            w.writerow([team, team_month_totals.get(team, 0), team_year_totals.get(team, 0)])
        w.writerow([])
        w.writerow(["人员岗位日期明细"])
        w.writerow(["姓名", "双流主班", "双流副班(空勤)", "双流副班(地勤)", "天府主班", "天府副班", "机队总负责"])
        for e in employees:
            row = [e.name]
            for pos in POSITION_EXPORT_ORDER:
                row.append("、".join(person_dates[e.id].get(pos, [])))
            w.writerow(row)


def export_excel(
    path: Path,
    assignments: list[Assignment],
    employees: list[Employee],
    day_tags: dict | None = None,
    person_year_totals: dict[int, int] | None = None,
    team_year_totals: dict[str, int] | None = None,
) -> None:
    by_id = {e.id: e for e in employees}
    day_tags = day_tags or {}
    person_year_totals = person_year_totals or {}
    team_year_totals = team_year_totals or {}
    wb = Workbook()

    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="1F2937")
    detail_font = Font(size=12)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = "排班结果"
    ws.merge_cells("A1:K1")
    ws["A1"] = "排班结果明细"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.append([])
    ws.append(["日期", "星期", "是否周末/节假日", "双流主班", "双流副班（空勤）", "双流副班（地勤）", "天府主班", "天府副班", "机队总负责", "双流主班大队", "天府主班大队"])

    weekday_map = {0: "星期一", 1: "星期二", 2: "星期三", 3: "星期四", 4: "星期五", 5: "星期六", 6: "星期日"}
    by_date_pos: dict = {}
    person_dates = _person_position_dates(assignments)
    for a in assignments:
        by_date_pos[(a.work_date, a.position)] = a

    all_dates = sorted({a.work_date for a in assignments})
    if all_dates:
        d = all_dates[0]
        end = all_dates[-1]
        continuous_dates = []
        while d <= end:
            continuous_dates.append(d)
            d += timedelta(days=1)
    else:
        continuous_dates = []

    def is_special(d):
        if d in day_tags:
            return day_tags[d] == "SPECIAL"
        return d.weekday() >= 5

    def name_of(d, pos):
        a = by_date_pos.get((d, pos))
        if not a:
            return ""
        return by_id[a.employee_id].name

    def team_of(d, pos):
        a = by_date_pos.get((d, pos))
        if not a:
            return ""
        return by_id[a.employee_id].team

    for d in continuous_dates:
        ws.append([
            d.isoformat(),
            weekday_map[d.weekday()],
            "是" if is_special(d) else "否",
            name_of(d, "SL_MAIN"),
            name_of(d, "SL_AIR"),
            name_of(d, "SL_GROUND"),
            name_of(d, "TF_MAIN"),
            name_of(d, "TF_GROUND"),
            name_of(d, "FLEET_LEAD"),
            team_of(d, "SL_MAIN"),
            team_of(d, "TF_MAIN"),
        ])

    for c, w in zip("ABCDEFGHIJK", [14, 10, 14, 16, 18, 18, 16, 14, 16, 14, 14]):
        ws.column_dimensions[c].width = w
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = border
            cell.alignment = center
            if cell.row == 3:
                cell.fill = head_fill
                cell.font = head_font

    person_month = wb.create_sheet("人员月度统计")
    person_month.merge_cells("A1:E1")
    person_month["A1"] = "人员月度统计"
    person_month["A1"].font = title_font
    person_month["A1"].alignment = center
    person_month.append([])
    person_month.append(["姓名", "大队", "中队", "月度值班总数", "年度值班总数"])
    p_counter = defaultdict(int)
    for a in assignments:
        p_counter[a.employee_id] += 1
    for e in employees:
        person_month.append([e.name, e.team, e.squad, p_counter[e.id], person_year_totals.get(e.id, 0)])

    for c, w in zip("ABCDE", [14, 10, 12, 14, 14]):
        person_month.column_dimensions[c].width = w
    for row in person_month.iter_rows(min_row=3, max_row=person_month.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            cell.alignment = center
            if cell.row == 3:
                cell.fill = head_fill
                cell.font = head_font

    team_month = wb.create_sheet("大队月度统计")
    team_month.merge_cells("A1:C1")
    team_month["A1"] = "大队月度统计"
    team_month["A1"].font = title_font
    team_month["A1"].alignment = center
    team_month.append([])
    team_month.append(["大队", "月度值班总数", "年度值班总数"])
    t_counter = defaultdict(int)
    for a in assignments:
        t_counter[by_id[a.employee_id].team] += 1
    for team in sorted({e.team for e in employees}):
        team_month.append([team, t_counter.get(team, 0), team_year_totals.get(team, 0)])

    for c, w in zip("ABC", [12, 16, 16]):
        team_month.column_dimensions[c].width = w
    for row in team_month.iter_rows(min_row=3, max_row=team_month.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            cell.alignment = center
            if cell.row == 3:
                cell.fill = head_fill
                cell.font = head_font

    person_year = wb.create_sheet("人员年度统计")
    person_year.merge_cells("A1:D1")
    person_year["A1"] = "人员年度统计"
    person_year["A1"].font = title_font
    person_year["A1"].alignment = center
    person_year.append([])
    person_year.append(["姓名", "大队", "中队", "年度值班总数"])
    for e in employees:
        person_year.append([e.name, e.team, e.squad, person_year_totals.get(e.id, 0)])
    for c, w in zip("ABCD", [14, 10, 12, 16]):
        person_year.column_dimensions[c].width = w
    for row in person_year.iter_rows(min_row=3, max_row=person_year.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = center
            if cell.row == 3:
                cell.fill = head_fill
                cell.font = head_font

    team_year = wb.create_sheet("大队年度统计")
    team_year.merge_cells("A1:B1")
    team_year["A1"] = "大队年度统计"
    team_year["A1"].font = title_font
    team_year["A1"].alignment = center
    team_year.append([])
    team_year.append(["大队", "年度值班总数"])
    for team in sorted({e.team for e in employees}):
        team_year.append([team, team_year_totals.get(team, 0)])
    for c, w in zip("AB", [12, 16]):
        team_year.column_dimensions[c].width = w
    for row in team_year.iter_rows(min_row=3, max_row=team_year.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = center
            if cell.row == 3:
                cell.fill = head_fill
                cell.font = head_font

    person_position = wb.create_sheet("人员岗位日期")
    person_position.merge_cells("A1:G1")
    person_position["A1"] = "人员岗位日期明细"
    person_position["A1"].font = title_font
    person_position["A1"].alignment = center
    person_position.append([])
    person_position.append(["姓名", "双流主班", "双流副班(空勤)", "双流副班(地勤)", "天府主班", "天府副班", "机队总负责"])
    for e in employees:
        row = [e.name]
        max_lines = 1
        for pos in POSITION_EXPORT_ORDER:
            text = "\n".join(person_dates[e.id].get(pos, []))
            row.append(text)
            max_lines = max(max_lines, len(person_dates[e.id].get(pos, [])) or 1)
        person_position.append(row)
        person_position.row_dimensions[person_position.max_row].height = max(28, 24 * max_lines)

    for c, w in zip("ABCDEFG", [14, 18, 20, 20, 18, 16, 18]):
        person_position.column_dimensions[c].width = w
    for row in person_position.iter_rows(min_row=3, max_row=person_position.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row == 3:
                cell.fill = head_fill
                cell.font = head_font
            else:
                cell.font = detail_font

    wb.save(path)
